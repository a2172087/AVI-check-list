import sys
import os
import logging
from PyQt5.QtWidgets import QApplication, QWidget, QPushButton, QVBoxLayout, QFileDialog, QProgressBar, QMessageBox, QLabel, QDesktopWidget
from PyQt5.QtCore import QThread, pyqtSignal, Qt
from PyQt5.QtGui import QFont, QIcon
import configparser
import re
import qtmodern.styles
import qtmodern.windows
import json
import shutil
from openpyxl import load_workbook
import openpyxl
import traceback
from PyQt5.QtGui import QPixmap
import py7zr
import datetime
import socket
import tempfile

logging.basicConfig(level=logging.ERROR, format='%(asctime)s - %(levelname)s - %(message)s')

def resource_path(relative_path):
    try:
        base_path = sys._MEIPASS
    except Exception:
        base_path = os.path.abspath(".")

    return os.path.join(base_path, relative_path)

class FileProcessor(QThread):
    progress_updated = pyqtSignal(int)
    processing_completed = pyqtSignal()
    error_occurred = pyqtSignal(str)
    open_folder_signal = pyqtSignal(str) 

    def __init__(self, avi_recipe_path):
        super().__init__()
        self.avi_recipe_path = avi_recipe_path
        self.variables = {'Default': {}, 'Default1': {}}
        self.default1_name = ''
        self.default1_actual_name = '' 
        self.surface_on_sb_variables = {}
        self.uniform_surface_on_sb_variables = {}
        
        # 提取 'Recipe/' 之後的部分作為 AVI_recipe_name
        recipe_index = avi_recipe_path.rfind('Recipe/')
        if recipe_index != -1:
            self.variables['AVI_recipe_name'] = avi_recipe_path[recipe_index + 7:]  # 7 是 'Recipe/' 的長度
        else:
            self.variables['AVI_recipe_name'] = os.path.basename(avi_recipe_path)
        
        # 解析 AVI_recipe_group_ID 和 AVI_recipe_EQP_ID
        recipe_parts = self.variables['AVI_recipe_name'].split('-')
        if len(recipe_parts) == 5:
            self.variables['AVI_recipe_EQP_ID'] = recipe_parts[0]
            self.variables['AVI_recipe_group_ID'] = recipe_parts[1]
        else:
            raise ValueError('Recipe檔名錯誤，請遵照EQP-Group-Stage-E-Version的格式進行命名')

    def clean_text(self, text):
        return ''.join(char for char in text if ord(char) < 128)

    def run(self):
        try:
            self.process_files()
            self.update_excel_file()
            
            print("Result：")
            print(json.dumps(self.variables, indent=2))
            
            self.processing_completed.emit()
        except Exception as e:
            self.error_occurred.emit(str(e))

    def process_files(self):
        # 首先處理 WaferMapRecipe.ini
        setup1_path = os.path.join(self.avi_recipe_path, 'Setup1')
        if os.path.exists(setup1_path):
            wafer_map_recipe_path = os.path.join(setup1_path, 'WaferMapRecipe.ini')
            if os.path.exists(wafer_map_recipe_path):
                self.parse_wafer_map_recipe(wafer_map_recipe_path)
            else:
                print("警告: 在 Setup1 資料夾中未找到 WaferMapRecipe.ini 文件")
        else:
            print("警告: 未找到 Setup1 資料夾")

        # 繼續處理其他文件
        setup1_path = os.path.join(self.avi_recipe_path, 'Setup1')
        recipes_path = os.path.join(setup1_path, 'Recipes')
        
        print(f"Recipes path: {recipes_path}")
        
        # 處理 Default 資料夾
        default_path = os.path.join(recipes_path, 'Default')
        print(f"Processing Default folder: {default_path}")
        self.process_folder(default_path, 'Default', 0)  # 從 0 開始計數
        
        # 尋找其他資料夾（可能的 Default1）
        other_folders = [f for f in os.listdir(recipes_path) if f != 'Default' and os.path.isdir(os.path.join(recipes_path, f))]
        print(f"Other folders found: {other_folders}")
        
        # 根據 other_folders 的數量設置 Recipe_file_count
        self.Recipe_file_count = 'Multi' if len(other_folders) >= 1 else 'Single'
        self.variables['Recipe_file_count'] = 'Multi' if len(other_folders) >= 1 else 'Single'
        print(f"Recipe_file_count: {self.Recipe_file_count}")
        
        if len(other_folders) >= 2:  # 如果有 2 個或更多額外的資料夾（不包括 Default）
            raise Exception(f"Setup1\\Recipes\\file count >={len(other_folders) + 1}, 請使用者檢查Recipe的數量|{recipes_path}")
        
        if other_folders:
            default1_path = os.path.join(recipes_path, other_folders[0])
            self.default1_actual_name = os.path.basename(default1_path)
            self.default1_name = 'Default1'
            print(f"Default1 folder actual name: {self.default1_actual_name}")
            print(f"Processing Default1 folder: {default1_path}")
            self.process_folder(default1_path, 'Default1', 0)  # 重新從 0 開始計數
        else:
            print("No Default1 folder found")

    def process_folder(self, folder_path, folder_type, initial_bump_map_count):
        print(f"Entering process_folder for {folder_type}: {folder_path}")
        self.bump_map_count = initial_bump_map_count
        self.zone_to_bump_map = {}  # 重置映射
        
        files_to_process = [
            ('OpticsPreset.ini', self.parse_optics_preset),
            ('AlignRtp.ini', self.parse_align_rtp),
            ('ProductInfo.ini', self.parse_product_info),
            ('AlignmentData.ini', self.parse_alignment_data),
            ('Recipe.ini', self.parse_recipe),
            ('RTP.txt', self.parse_rtp)
        ]

        for filename, parse_function in files_to_process:
            file_path = self.find_file(filename, folder_path)
            if file_path:
                print(f"Found and processing {filename} in {folder_type}")
                parse_function(file_path, folder_type)
            else:
                print(f"File not found: {filename} in {folder_type}")

        print(f"Finished processing {folder_type}, found {self.bump_map_count} Bump Maps")

        # 列出 Zones 資料夾中的所有文件
        zones_path = os.path.join(folder_path, 'Zones')
        if os.path.exists(zones_path):
            print(f"Files in {folder_type} Zones folder:")
            for file in os.listdir(zones_path):
                print(f"  - {file}")
        else:
            print(f"Zones folder not found in {folder_type}")

    def find_file(self, filename, search_path):
        for root, dirs, files in os.walk(search_path):
            if filename in files:
                return os.path.join(root, filename)
        return None

    def parse_optics_preset(self, file_path, folder_type):
        config = configparser.ConfigParser()
        config.optionxform = str  # 保持鍵的大小寫
        with open(file_path, 'r', encoding='utf-8', errors='ignore') as file:
            config.read_file(file)

        if 'RobotSetup' in config:
            robotsetup_name = config['RobotSetup'].get('Name', '')
            if robotsetup_name:
                self.variables[folder_type]['OpticsPreset_Robotsetup_Name'] = robotsetup_name

        if 'General' in config:
            general = config['General']

            scan2d_mag = next((v for k, v in general.items() if k.startswith('Scan2d-Mag')), None)
            if scan2d_mag:
                self.variables[folder_type]['OpticsPreset_General_Scan2d_Mag'] = scan2d_mag

            verify_color_mag = next((v for k, v in general.items() if 'VerifyColorMag' in k and k.endswith('-Mag')), None)
            if verify_color_mag:
                self.variables[folder_type]['OpticsPreset_General_VerifyColorMag_Mag'] = verify_color_mag

            diff_light = next((v for k, v in general.items() if k.startswith('DiffLight') and '-' not in k), None)
            if diff_light:
                self.variables[folder_type]['OpticsPreset_General_DiffLight'] = self.round_to_one_decimal(diff_light)

            ref_light = next((v for k, v in general.items() if k.startswith('RefLight') and '-' not in k), None)
            if ref_light:
                self.variables[folder_type]['OpticsPreset_General_RefLight'] = self.round_to_one_decimal(ref_light)

            verify_color_ref_light = next((v for k, v in general.items() if 'VerifyColorMag' in k and k.endswith('-RefLight')), None)
            if verify_color_ref_light:
                self.variables[folder_type]['OpticsPreset_General_VerifyColorMag_RefLight'] = self.round_to_one_decimal(verify_color_ref_light)

    def round_to_one_decimal(self, value):
        try:
            return f"{float(value):.1f}"
        except ValueError:
            return value

    def parse_wafer_map_recipe(self, file_path):
        config = configparser.ConfigParser()
        with open(file_path, 'r', encoding='utf-8', errors='ignore') as file:
            content = file.read()
        cleaned_content = self.clean_text(content)
        config.read_string(cleaned_content)

        self.variables['WaferMapRecipe_GENERAL_ExportInAutoCycle'] = self.clean_text(config.get('GENERAL', 'ExportInAutoCycle', fallback=''))
        self.variables['WaferMapRecipe_Input_Update_Enable'] = self.clean_text(config.get('Input_Update', 'Enable', fallback=''))
        self.variables['WaferMapRecipe_Input_Update_FileMask'] = self.clean_text(config.get('Input_Update', 'FileMask', fallback=''))
        self.variables['WaferMapRecipe_Input_Update_ImportDirectory'] = self.clean_text(config.get('Input_Update', 'ImportDirectory', fallback=''))
        self.variables['WaferMapRecipe_Input_Update_ConverterName'] = self.clean_text(config.get('Input_Update', 'ConverterName', fallback=''))

    def parse_align_rtp(self, file_path, folder_type):
        config = configparser.ConfigParser()
        with open(file_path, 'r', encoding='utf-8', errors='ignore') as file:
            content = file.read()
        cleaned_content = self.clean_text(content)
        config.read_string(cleaned_content)

        self.variables[folder_type]['AlignRtp_DIE_Alignment_Die__MinScore'] = self.clean_text(config.get('DIE Alignment', 'Die__MinScore', fallback=''))

    def parse_product_info(self, file_path, folder_type):
        config = configparser.ConfigParser()
        with open(file_path, 'r', encoding='utf-8', errors='ignore') as file:
            content = file.read()
        cleaned_content = self.clean_text(content)
        config.read_string(cleaned_content)

        self.variables[folder_type]['ProductInfo_General_OCRWaferIDMask'] = self.clean_text(config.get('General', 'OCRWaferIDMask', fallback=''))
        self.variables[folder_type]['ProductInfo_Geometric_XDieIndex'] = self.clean_text(config.get('Geometric', 'XDieIndex', fallback=''))
        self.variables[folder_type]['ProductInfo_Geometric_YDieIndex'] = self.clean_text(config.get('Geometric', 'YDieIndex', fallback=''))
        self.variables[folder_type]['ProductInfo_Geometric_Diameter'] = self.clean_text(config.get('Geometric', 'Diameter', fallback=''))
        self.variables[folder_type]['ProductInfo_UpperIdReader_Enabled'] = self.clean_text(config.get('UpperIdReader', 'Enabled', fallback=''))
        self.variables[folder_type]['ProductInfo_UpperIdReader_JobName'] = self.clean_text(config.get('UpperIdReader', 'JobName', fallback=''))

    def parse_alignment_data(self, file_path, folder_type):
        config = configparser.ConfigParser()
        with open(file_path, 'r', encoding='utf-8', errors='ignore') as file:
            content = file.read()
        cleaned_content = self.clean_text(content)
        config.read_string(cleaned_content)

        self.variables[folder_type]['AlignmentData_General_MinScore'] = self.clean_text(config.get('General', 'MinScore', fallback=''))

    def parse_recipe(self, file_path, folder_type):
        config = configparser.ConfigParser()
        with open(file_path, 'r', encoding='utf-8', errors='ignore') as file:
            content = file.read()
        cleaned_content = self.clean_text(content)
        config.read_string(cleaned_content)

        self.variables[folder_type]['Recipe_AutoCycle_ExportPMdata'] = self.clean_text(config.get('AutoCycle', 'ExportPMdata', fallback=''))
        self.variables[folder_type]['Recipe_AutoCycle_MaxImagesToGrabDie'] = self.clean_text(config.get('AutoCycle', 'MaxImagesToGrabDie', fallback=''))

    def parse_rtp(self, file_path, folder_type):
        logging.info(f"Starting parse_rtp for folder_type: {folder_type}")
        logging.info(f"self.avi_recipe_path: {self.avi_recipe_path}")
        logging.info(f"Parsing RTP file: {file_path}")
        
        try:
            with open(file_path, 'r', encoding='utf-8', errors='ignore') as file:
                content = file.read()
        except IOError as e:
            logging.error(f"Error reading file {file_path}: {e}")
            return

        cleaned_content = self.clean_text(content)
        sections = re.split(r'(\[.*?\].*?\n)', cleaned_content)
        
        result_sections = []
        zone_status = {}
        bump_map_count = 0  # 重置計數
        zone_to_bump_map = {}

        print(f"\n--- {folder_type} Zones ---")
        for section in sections:
            if section.strip().startswith('[') and section.strip().endswith('; Zone name'):
                zone_name = re.search(r'\[(.*?)\]', section).group(1)
                if zone_name not in ['PostProcess', 'Scan_Area']:
                    bump_map_count += 1
                    if bump_map_count <= 5:
                        zone_to_bump_map[zone_name] = f'Bump_Map_{bump_map_count}'

        logging.info(f"Identified zones: {zone_to_bump_map}")

        # 定義不區分大小寫的文件查找函數
        def find_file_case_insensitive(dir_path, filename):
            for file in os.listdir(dir_path):
                if file.lower() == filename.lower():
                    return os.path.join(dir_path, file)
            return None

        # 分析所有區域的狀態
        for zone_name, bump_map_name in zone_to_bump_map.items():
            normalized_zone_name = zone_name.replace('_', ' ')
            actual_folder_type = self.default1_actual_name if folder_type == 'Default1' else folder_type
            zones_dir = os.path.join(self.avi_recipe_path, 'Setup1', 'Recipes', actual_folder_type, 'Zones')
            ini_file = os.path.join(zones_dir, f'{normalized_zone_name}.ini')
            
            logging.info(f"Processing zone: {zone_name} as {bump_map_name} for {actual_folder_type}")
            logging.info(f"Looking for INI file: {ini_file}")
            
            # 使用不區分大小寫的文件查找
            found_ini_file = find_file_case_insensitive(zones_dir, f'{normalized_zone_name}.ini')
            
            if found_ini_file:
                print(f"Found INI file: {os.path.basename(found_ini_file)}")
                logging.info(f"INI file exists: {found_ini_file}")
                zone_status[bump_map_name] = {}
                config = configparser.ConfigParser()
                config.read(found_ini_file)
                for alg in ['Solder Bump', 'Surface on SB', 'Uniform Surface on SB', 'Surface', 'PMI Advanced', 'Probe Mark Inspection']:
                    zone_status[bump_map_name][alg] = config.getboolean(alg, 'Enable', fallback=False)
            else:
                print(f"INI file not found for: {normalized_zone_name}.ini")
                logging.warning(f"INI file not found for {zone_name} in {actual_folder_type}. Assuming all algorithms are disabled.")
                # 列出目標目錄中的所有文件
                logging.info(f"Files in {zones_dir}:")
                for file in os.listdir(zones_dir):
                    logging.info(f"  - {file}")
                zone_status[bump_map_name] = {alg: False for alg in ['Solder Bump', 'Surface on SB', 'Uniform Surface on SB', 'Surface', 'PMI Advanced', 'Probe Mark Inspection']}
            
            logging.info(f"Zone status for {bump_map_name} in {actual_folder_type}: {zone_status[bump_map_name]}")

        # 處理每個區域
        for section in sections:
            if section.strip().startswith('[') and section.strip().endswith('; Zone name'):
                zone_name = re.search(r'\[(.*?)\]', section).group(1)
                if zone_name in ['PostProcess', 'Scan_Area']:
                    result_sections.append(section)
                elif zone_name in zone_to_bump_map:
                    bump_map_name = zone_to_bump_map[zone_name]
                    if any(zone_status.get(bump_map_name, {}).values()):
                        new_section = f'[{bump_map_name}]   ; Zone name\n'
                        logging.info(f"Converted {zone_name} to {new_section.strip()} for {actual_folder_type}")
                    else:
                        new_section = '[Fail]   ; Zone name\n'
                        logging.info(f"Marked {zone_name} as Fail for {actual_folder_type}")
                    result_sections.append(new_section)
            else:
                result_sections.append(section)

        new_content = ''.join(result_sections)

        # 處理 Bump Map 部分
        for bump_map_name, status in zone_status.items():
            bump_map_sections = re.findall(f'\[{bump_map_name}\].*?(?=\[Bump_Map|\[Fail|\[Scan_Area|\Z)', new_content, re.DOTALL)
            for section in bump_map_sections:
                alg_sections = re.split(r'\nAlg\s*=\s*', section)
                for alg_section in alg_sections[1:]:
                    alg_type = alg_section.split('\n')[0].strip()
                    alg_type_normalized = alg_type.replace('_', ' ')
                    if status.get(alg_type_normalized, False):
                        prefix = f'RTP_{bump_map_name}_{alg_type}'
                        logging.info(f"Parsing section for {prefix} in {actual_folder_type}")
                        self.parse_section(alg_section, prefix, folder_type)
                    else:
                        logging.warning(f"Skipping disabled algorithm {alg_type} for {bump_map_name} in {actual_folder_type}")

        # 處理 Scan Area 部分
        scan_area_surface = re.search(r'\[Scan_Area\].*?(?=\[|\Z)', new_content, re.DOTALL)
        if scan_area_surface:
            logging.info(f"Parsing Scan Area Surface section for {actual_folder_type}")
            self.parse_section(scan_area_surface.group(), 'RTP_Scan_Area_Surface', folder_type)
        else:
            logging.warning(f"Scan Area Surface section not found for {actual_folder_type}")

        # 處理 Fail 部分
        fail_sections = re.findall(r'\[Fail\].*?(?=\[Bump_Map|\[Scan_Area|\Z)', new_content, re.DOTALL)
        for section in fail_sections:
            logging.info(f"Parsing Fail section for {actual_folder_type}, but not parsing it")

        logging.info(f"Parsed data for {actual_folder_type}: {self.variables.get(folder_type, {})}")

    def parse_section(self, section_content, prefix, folder_type):
        lines = section_content.split('\n')
        for line in lines[1:]: 
            if '=' in line:
                key, value = line.split('=', 1)
                key = self.clean_text(key.strip())
                key = key.replace('[', '').replace(']', '')
                value = self.clean_text(value.split(';')[0].strip())
                if value.startswith('.'):
                    value = '0' + value
                self.variables.setdefault(folder_type, {})[f"{prefix}_{key}"] = value

        self.variables.setdefault(folder_type, {})[f"{prefix}_Alg"] = prefix.split('_')[-1]

    def parse_uniform_surface_on_sb(self, section_content, bump_map_number):
        allowed_params = [
            'Enable_Moving_Surface', 'Exposed_Area_High_TH', 'Exposed_Area_Low_TH',
            'Position_Don\'t-Care_Width', 'Original_position_don\'t_care_width',
            'Min_Defect_Area_-_Bright', 'Min_Defect_Width_-_Bright',
            'Min_Defect_Length_-_Bright', 'Contrast_Upper_value_-_Bright',
            'Min_Defect_Area_-_Dark', 'Min_Defect_Width_-_Dark',
            'Min_Defect_Length_-_Dark', 'Contrast_Lower_value_-_Dark',
            'MaxAreaSum', 'CollectForGlobalSum'
        ]

        lines = section_content.split('\n')
        for line in lines:
            if '=' in line:
                key, value = line.split('=', 1)
                key = self.clean_text(key.strip())
                if key in allowed_params:
                    value = self.clean_text(value.split(';')[0].strip())
                    if value.startswith('.'):
                        value = '0' + value
                    self.uniform_surface_on_sb_variables[f"RTP_Bump_Map_{bump_map_number}_Uniform_Surface_on_SB_{key}"] = value

    def update_excel_file(self):
        template_path = r"D:\本地應用程式\AVI Check list\Camtek Falcon Check list_V4.xlsx"
        downloads_folder = os.path.join(os.path.expanduser("~"), "Downloads")
        # 從 avi_recipe_path 提取檔案名稱
        recipe_name = os.path.basename(self.avi_recipe_path)
        # 創建新的檔案名稱
        new_file_name = f"{recipe_name}_AVI check list.xlsx"
        # 組合完整的輸出路徑
        output_path = os.path.join(downloads_folder, new_file_name)

        # Copy the template file to the Downloads folder
        shutil.copy2(template_path, output_path)

        try:
            # Open the copied file
            wb = load_workbook(output_path)

            # 解鎖所有工作表
            for ws in wb.worksheets:
                if ws.protection.sheet:
                    ws.protection.password = 'Ardentec'
                    ws.protection.enable()
                    ws.protection.disable()  # 解除保護
            
            # Define mappings
            check_list_mappings = {
                'AVI_recipe_group_ID': 'C4',
                'ProductInfo_Geometric_Diameter': 'C5',
                'ProductInfo_Geometric_XDieIndex': 'C7',
                'ProductInfo_Geometric_YDieIndex': 'C8',
                'AVI_recipe_name': 'C16',
                'AVI_recipe_EQP_ID': 'C17',
                'AlignRtp_DIE_Alignment_Die__MinScore': 'C18',
                'AlignmentData_General_MinScore': 'C19',  
                'ProductInfo_UpperIdReader_JobName': 'C20',  
                'OpticsPreset_Robotsetup_Name': 'C21',
                'OpticsPreset_General_DiffLight': 'C23',
                'OpticsPreset_General_RefLight': 'C24',
                'OpticsPreset_General_VerifyColorMag_RefLight': 'C25',
                'OpticsPreset_General_VerifyColorMag_Mag': 'C31',
                'Recipe_AutoCycle_ExportPMdata': 'C35',
                'Recipe_AutoCycle_MaxImagesToGrabDie': 'C36',
                'ProductInfo_General_OCRWaferIDMask': 'C37',
                'Recipe_file_count': 'C38',
                'WaferMapRecipe_GENERAL_ExportInAutoCycle': 'C40',
                'WaferMapRecipe_Input_Update_Enable': 'C42',
                'WaferMapRecipe_Input_Update_FileMask': 'C43',
                'WaferMapRecipe_Input_Update_ImportDirectory': 'C44',
                'WaferMapRecipe_Input_Update_ConverterName': 'C45',
            }

            surface_mappings = {
                'RTP_Scan_Area_Surface_Min_Defect_Area_-_Bright': 'F4',
                'RTP_Scan_Area_Surface_Min_Defect_Width_-_Bright': 'F5',
                'RTP_Scan_Area_Surface_Min_Defect_Length_-_Bright': 'F6',
                'RTP_Scan_Area_Surface_Contrast_Delta_-_Bright': 'F7',
                'RTP_Scan_Area_Surface_Contrast_Factor_-_Bright': 'F8',
                'RTP_Scan_Area_Surface_Min_Defect_Area_-_Dark': 'F9',
                'RTP_Scan_Area_Surface_Min_Defect_Width_-_Dark': 'F10',
                'RTP_Scan_Area_Surface_Min_Defect_Length_-_Dark': 'F11',
                'RTP_Scan_Area_Surface_Contrast_Delta_-_Dark': 'F12',
                'RTP_Scan_Area_Surface_Contrast_Factor_-_Dark': 'F13',
                'RTP_Scan_Area_Surface_Big_Area_Status_-_Bright': 'F14',
                'RTP_Scan_Area_Surface_Big_Area_Status_-_Dark': 'F15',
                'RTP_Scan_Area_Surface_Cluster_Area': 'F16',
                'RTP_Scan_Area_Surface_Cluster_Distance': 'F17',
                'RTP_Scan_Area_Surface_Cluster_Diameter': 'F18',
                'RTP_Scan_Area_Surface_Adaptive_Histogram_Mode': 'F19',
                'RTP_Scan_Area_Surface_CollectForGlobalSum': 'F20',
                'RTP_Scan_Area_Surface_MaxAreaSum': 'F21',
                'RTP_Scan_Area_Surface_Zone_CD_Radius': 'F22',
                'RTP_Scan_Area_Surface_Dark_Zone_CD_Percent': 'F23',
                'RTP_Scan_Area_Surface_Bright_Zone_CD_Percent': 'F24',
                'RTP_Scan_Area_Surface_MaxCountSum': 'F25'
                }

            pad_device_mappings = {
                #[Bump_Map_1]
                'RTP_Bump_Map_1_Surface_Min_Defect_Area_-_Bright': 'F4',
                'RTP_Bump_Map_1_Surface_Min_Defect_Width_-_Bright': 'F5',
                'RTP_Bump_Map_1_Surface_Min_Defect_Length_-_Bright': 'F6',
                'RTP_Bump_Map_1_Surface_Contrast_Delta_-_Bright': 'F7',
                'RTP_Bump_Map_1_Surface_Contrast_Factor_-_Bright': 'F8',
                'RTP_Bump_Map_1_Surface_Min_Defect_Area_-_Dark': 'F9',
                'RTP_Bump_Map_1_Surface_Min_Defect_Width_-_Dark': 'F10',
                'RTP_Bump_Map_1_Surface_Min_Defect_Length_-_Dark': 'F11',
                'RTP_Bump_Map_1_Surface_Contrast_Delta_-_Dark': 'F12',
                'RTP_Bump_Map_1_Surface_Contrast_Factor_-_Dark': 'F13',
                'RTP_Bump_Map_1_Surface_Big_Area_Status_-_Bright': 'F14',
                'RTP_Bump_Map_1_Surface_Big_Area_Status_-_Dark': 'F15',
                'RTP_Bump_Map_1_Surface_Cluster_Area': 'F16',
                'RTP_Bump_Map_1_Surface_Cluster_Distance': 'F17',
                'RTP_Bump_Map_1_Surface_Cluster_Diameter': 'F18',
                'RTP_Bump_Map_1_Surface_Adaptive_Histogram_Mode': 'F19',
                'RTP_Bump_Map_1_Surface_CollectForGlobalSum': 'F20',
                'RTP_Bump_Map_1_Surface_MaxAreaSum': 'F21',
                'RTP_Bump_Map_1_Surface_Zone_CD_Radius': 'F22',
                'RTP_Bump_Map_1_Surface_Dark_Zone_CD_Percent': 'F23',
                'RTP_Bump_Map_1_Surface_Bright_Zone_CD_Percent': 'F24',
                'RTP_Bump_Map_1_Surface_MaxCountSum': 'F25',
                'RTP_Bump_Map_1_Probe_Mark_Inspection_Pad_Is_Rectangle': 'F125',
                'RTP_Bump_Map_1_Probe_Mark_Inspection_USL_Pad_Size_[X]': 'F126',
                'RTP_Bump_Map_1_Probe_Mark_Inspection_LSL_Pad_Size_[X]': 'F127',
                'RTP_Bump_Map_1_Probe_Mark_Inspection_USL_Pad_Size_[Y]': 'F128',
                'RTP_Bump_Map_1_Probe_Mark_Inspection_LSL_Pad_Size_[Y]': 'F129',
                'RTP_Bump_Map_1_Probe_Mark_Inspection_Pad_Mislocation_[X]': 'F130',
                'RTP_Bump_Map_1_Probe_Mark_Inspection_Pad_Mislocation_[Y]': 'F131',
                'RTP_Bump_Map_1_Probe_Mark_Inspection_Pad_Sensitivity': 'F132',
                'RTP_Bump_Map_1_Probe_Mark_Inspection_Probe__Sensitivity': 'F133',
                'RTP_Bump_Map_1_Probe_Mark_Inspection_Pad_Low_Threshold': 'F134',
                'RTP_Bump_Map_1_Probe_Mark_Inspection_Pad_High_Threshold': 'F135',
                'RTP_Bump_Map_1_Probe_Mark_Inspection_Max_Area_For_Noise_[Spots]': 'F136',
                'RTP_Bump_Map_1_Probe_Mark_Inspection_PM_Max_Area_[%_From_pad]': 'F137',
                'RTP_Bump_Map_1_Probe_Mark_Inspection_PM_Min_Area': 'F138',
                'RTP_Bump_Map_1_Probe_Mark_Inspection_Max_Number_Of_Prob_Marks': 'F139',
                'RTP_Bump_Map_1_Probe_Mark_Inspection_Min_Number_Of_Prob_Marks': 'F140',
                'RTP_Bump_Map_1_Probe_Mark_Inspection_Min_acceptable_distance__from_Pad': 'F141',
                'RTP_Bump_Map_1_Probe_Mark_Inspection_Max_PM_size_allowed_touching_the_Pad': 'F142',
                'RTP_Bump_Map_1_Probe_Mark_Inspection_Enable_surface_zone': 'F143',
                'RTP_Bump_Map_1_Probe_Mark_Inspection_Dont_Care_zone': 'F144',
                'RTP_Bump_Map_1_Probe_Mark_Inspection_Surface_Zone': 'F145',
                'RTP_Bump_Map_1_Probe_Mark_Inspection_Defect_Area_Inside_Surface': 'F146',
                'RTP_Bump_Map_1_Probe_Mark_Inspection_Contrast_Delta_-_Dark': 'F147',
                'RTP_Bump_Map_1_Probe_Mark_Inspection_Contrast_Delta_-_bright': 'F148',
                'RTP_Bump_Map_1_PMI_Advanced_USL_Pad_Size_X': 'F256',
                'RTP_Bump_Map_1_PMI_Advanced_LSL_Pad_Size_X': 'F257',
                'RTP_Bump_Map_1_PMI_Advanced_USL_Pad_Size_Y': 'F258',
                'RTP_Bump_Map_1_PMI_Advanced_LSL_Pad_Size_Y': 'F259',
                'RTP_Bump_Map_1_PMI_Advanced_Pad_Mislocation_X': 'F260',
                'RTP_Bump_Map_1_PMI_Advanced_Pad_Mislocation_Y': 'F261',
                'RTP_Bump_Map_1_PMI_Advanced_Pad_Edge_Sensitivity': 'F262',
                'RTP_Bump_Map_1_PMI_Advanced_Pad_Sensitivity': 'F263',
                'RTP_Bump_Map_1_PMI_Advanced_PM_Sensitivity': 'F264',
                'RTP_Bump_Map_1_PMI_Advanced_Pad_Gray_Level': 'F265',
                'RTP_Bump_Map_1_PMI_Advanced_Pad_Edge_Gray_Level': 'F266',
                'RTP_Bump_Map_1_PMI_Advanced_Surface_Gray_Level': 'F267',
                'RTP_Bump_Map_1_PMI_Advanced_USL_PM_Area_[%]': 'F268',
                'RTP_Bump_Map_1_PMI_Advanced_LSL_PM_Area': 'F269',
                'RTP_Bump_Map_1_PMI_Advanced_PM_Min_Spot_Area': 'F270',
                'RTP_Bump_Map_1_PMI_Advanced_Max_Number_Of_Prob_Marks': 'F271',
                'RTP_Bump_Map_1_PMI_Advanced_Min_Number_Of_Prob_Marks': 'F272',
                'RTP_Bump_Map_1_PMI_Advanced_Min_acceptable_distance__from_Pad': 'F273',
                'RTP_Bump_Map_1_PMI_Advanced_Max_PM_size_allowed_touching_the_Pad': 'F274',
                'RTP_Bump_Map_1_PMI_Advanced_Enable_surface_zone': 'F275',
                'RTP_Bump_Map_1_PMI_Advanced_Don**_Care_zone': 'F276',
                'RTP_Bump_Map_1_PMI_Advanced_Surface_Zone': 'F277',
                'RTP_Bump_Map_1_PMI_Advanced_Min_Defect_Area': 'F278',
                'RTP_Bump_Map_1_PMI_Advanced_Contrast_Delta_-_Dark': 'F279',
                'RTP_Bump_Map_1_PMI_Advanced_Contrast_Delta_-_bright': 'F280',
                'RTP_Bump_Map_1_PMI_Advanced_nspection_Sensitivity': 'F281',
                'RTP_Bump_Map_1_PMI_Advanced_Ref_Sensitivity': 'F282',

                #[Bump_Map_2]
                'RTP_Bump_Map_2_Surface_Min_Defect_Area_-_Bright': 'F28',
                'RTP_Bump_Map_2_Surface_Min_Defect_Width_-_Bright': 'F29',
                'RTP_Bump_Map_2_Surface_Min_Defect_Length_-_Bright': 'F30',
                'RTP_Bump_Map_2_Surface_Contrast_Delta_-_Bright': 'F31',
                'RTP_Bump_Map_2_Surface_Contrast_Factor_-_Bright': 'F32',
                'RTP_Bump_Map_2_Surface_Min_Defect_Area_-_Dark': 'F33',
                'RTP_Bump_Map_2_Surface_Min_Defect_Width_-_Dark': 'F34',
                'RTP_Bump_Map_2_Surface_Min_Defect_Length_-_Dark': 'F35',
                'RTP_Bump_Map_2_Surface_Contrast_Delta_-_Dark': 'F36',
                'RTP_Bump_Map_2_Surface_Contrast_Factor_-_Dark': 'F37',
                'RTP_Bump_Map_2_Surface_Big_Area_Status_-_Bright': 'F38',
                'RTP_Bump_Map_2_Surface_Big_Area_Status_-_Dark': 'F39',
                'RTP_Bump_Map_2_Surface_Cluster_Area': 'F40',
                'RTP_Bump_Map_2_Surface_Cluster_Distance': 'F41',
                'RTP_Bump_Map_2_Surface_Cluster_Diameter': 'F42',
                'RTP_Bump_Map_2_Surface_Adaptive_Histogram_Mode': 'F43',
                'RTP_Bump_Map_2_Surface_CollectForGlobalSum': 'F44',
                'RTP_Bump_Map_2_Surface_MaxAreaSum': 'F45',
                'RTP_Bump_Map_2_Surface_Zone_CD_Radius': 'F46',
                'RTP_Bump_Map_2_Surface_Dark_Zone_CD_Percent': 'F47',
                'RTP_Bump_Map_2_Surface_Bright_Zone_CD_Percent': 'F48',
                'RTP_Bump_Map_2_Surface_MaxCountSum': 'F49',
                'RTP_Bump_Map_2_Probe_Mark_Inspection_Pad_Is_Rectangle': 'F151',
                'RTP_Bump_Map_2_Probe_Mark_Inspection_USL_Pad_Size_[X]': 'F152',
                'RTP_Bump_Map_2_Probe_Mark_Inspection_LSL_Pad_Size_[X]': 'F153',
                'RTP_Bump_Map_2_Probe_Mark_Inspection_USL_Pad_Size_[Y]': 'F154',
                'RTP_Bump_Map_2_Probe_Mark_Inspection_LSL_Pad_Size_[Y]': 'F155',
                'RTP_Bump_Map_2_Probe_Mark_Inspection_Pad_Mislocation_[X]': 'F156',
                'RTP_Bump_Map_2_Probe_Mark_Inspection_Pad_Mislocation_[Y]': 'F157',
                'RTP_Bump_Map_2_Probe_Mark_Inspection_Pad_Sensitivity': 'F158',
                'RTP_Bump_Map_2_Probe_Mark_Inspection_Probe__Sensitivity': 'F159',
                'RTP_Bump_Map_2_Probe_Mark_Inspection_Pad_Low_Threshold': 'F160',
                'RTP_Bump_Map_2_Probe_Mark_Inspection_Pad_High_Threshold': 'F161',
                'RTP_Bump_Map_2_Probe_Mark_Inspection_Max_Area_For_Noise_[Spots]': 'F162',
                'RTP_Bump_Map_2_Probe_Mark_Inspection_PM_Max_Area_[%_From_pad]': 'F163',
                'RTP_Bump_Map_2_Probe_Mark_Inspection_PM_Min_Area': 'F164',
                'RTP_Bump_Map_2_Probe_Mark_Inspection_Max_Number_Of_Prob_Marks': 'F165',
                'RTP_Bump_Map_2_Probe_Mark_Inspection_Min_Number_Of_Prob_Marks': 'F166',
                'RTP_Bump_Map_2_Probe_Mark_Inspection_Min_acceptable_distance__from_Pad': 'F167',
                'RTP_Bump_Map_2_Probe_Mark_Inspection_Max_PM_size_allowed_touching_the_Pad': 'F168',
                'RTP_Bump_Map_2_Probe_Mark_Inspection_Enable_surface_zone': 'F169',
                'RTP_Bump_Map_2_Probe_Mark_Inspection_Dont_Care_zone': 'F170',
                'RTP_Bump_Map_2_Probe_Mark_Inspection_Surface_Zone': 'F171',
                'RTP_Bump_Map_2_Probe_Mark_Inspection_Defect_Area_Inside_Surface': 'F172',
                'RTP_Bump_Map_2_Probe_Mark_Inspection_Contrast_Delta_-_Dark': 'F173',
                'RTP_Bump_Map_2_Probe_Mark_Inspection_Contrast_Delta_-_bright': 'F174',
                'RTP_Bump_Map_2_PMI_Advanced_USL_Pad_Size_X': 'F285',
                'RTP_Bump_Map_2_PMI_Advanced_LSL_Pad_Size_X': 'F286',
                'RTP_Bump_Map_2_PMI_Advanced_USL_Pad_Size_Y': 'F287',
                'RTP_Bump_Map_2_PMI_Advanced_LSL_Pad_Size_Y': 'F288',
                'RTP_Bump_Map_2_PMI_Advanced_Pad_Mislocation_X': 'F289',
                'RTP_Bump_Map_2_PMI_Advanced_Pad_Mislocation_Y': 'F290',
                'RTP_Bump_Map_2_PMI_Advanced_Pad_Edge_Sensitivity': 'F291',
                'RTP_Bump_Map_2_PMI_Advanced_Pad_Sensitivity': 'F292',
                'RTP_Bump_Map_2_PMI_Advanced_PM_Sensitivity': 'F293',
                'RTP_Bump_Map_2_PMI_Advanced_Pad_Gray_Level': 'F294',
                'RTP_Bump_Map_2_PMI_Advanced_Pad_Edge_Gray_Level': 'F295',
                'RTP_Bump_Map_2_PMI_Advanced_Surface_Gray_Level': 'F296',
                'RTP_Bump_Map_2_PMI_Advanced_USL_PM_Area_[%]': 'F297',
                'RTP_Bump_Map_2_PMI_Advanced_LSL_PM_Area': 'F298',
                'RTP_Bump_Map_2_PMI_Advanced_PM_Min_Spot_Area': 'F299',
                'RTP_Bump_Map_2_PMI_Advanced_Max_Number_Of_Prob_Marks': 'F300',
                'RTP_Bump_Map_2_PMI_Advanced_Min_Number_Of_Prob_Marks': 'F301',
                'RTP_Bump_Map_2_PMI_Advanced_Min_acceptable_distance__from_Pad': 'F302',
                'RTP_Bump_Map_2_PMI_Advanced_Max_PM_size_allowed_touching_the_Pad': 'F303',
                'RTP_Bump_Map_2_PMI_Advanced_Enable_surface_zone': 'F304',
                'RTP_Bump_Map_2_PMI_Advanced_Don**_Care_zone': 'F305',
                'RTP_Bump_Map_2_PMI_Advanced_Surface_Zone': 'F306',
                'RTP_Bump_Map_2_PMI_Advanced_Min_Defect_Area': 'F307',
                'RTP_Bump_Map_2_PMI_Advanced_Contrast_Delta_-_Dark': 'F308',
                'RTP_Bump_Map_2_PMI_Advanced_Contrast_Delta_-_bright': 'F309',
                'RTP_Bump_Map_2_PMI_Advanced_nspection_Sensitivity': 'F310',
                'RTP_Bump_Map_2_PMI_Advanced_Ref_Sensitivity': 'F311',

                #[Bump_Map_3]
                'RTP_Bump_Map_3_Surface_Min_Defect_Area_-_Bright': 'F52',
                'RTP_Bump_Map_3_Surface_Min_Defect_Width_-_Bright': 'F53',
                'RTP_Bump_Map_3_Surface_Min_Defect_Length_-_Bright': 'F54',
                'RTP_Bump_Map_3_Surface_Contrast_Delta_-_Bright': 'F55',
                'RTP_Bump_Map_3_Surface_Contrast_Factor_-_Bright': 'F56',
                'RTP_Bump_Map_3_Surface_Min_Defect_Area_-_Dark': 'F57',
                'RTP_Bump_Map_3_Surface_Min_Defect_Width_-_Dark': 'F58',
                'RTP_Bump_Map_3_Surface_Min_Defect_Length_-_Dark': 'F59',
                'RTP_Bump_Map_3_Surface_Contrast_Delta_-_Dark': 'F60',
                'RTP_Bump_Map_3_Surface_Contrast_Factor_-_Dark': 'F61',
                'RTP_Bump_Map_3_Surface_Big_Area_Status_-_Bright': 'F62',
                'RTP_Bump_Map_3_Surface_Big_Area_Status_-_Dark': 'F63',
                'RTP_Bump_Map_3_Surface_Cluster_Area': 'F64',
                'RTP_Bump_Map_3_Surface_Cluster_Distance': 'F65',
                'RTP_Bump_Map_3_Surface_Cluster_Diameter': 'F66',
                'RTP_Bump_Map_3_Surface_Adaptive_Histogram_Mode': 'F67',
                'RTP_Bump_Map_3_Surface_CollectForGlobalSum': 'F68',
                'RTP_Bump_Map_3_Surface_MaxAreaSum': 'F69',
                'RTP_Bump_Map_3_Surface_Zone_CD_Radius': 'F70',
                'RTP_Bump_Map_3_Surface_Dark_Zone_CD_Percent': 'F71',
                'RTP_Bump_Map_3_Surface_Bright_Zone_CD_Percent': 'F72',
                'RTP_Bump_Map_3_Surface_MaxCountSum': 'F73',
                'RTP_Bump_Map_3_Probe_Mark_Inspection_Pad_Is_Rectangle': 'F177',
                'RTP_Bump_Map_3_Probe_Mark_Inspection_USL_Pad_Size_[X]': 'F178',
                'RTP_Bump_Map_3_Probe_Mark_Inspection_LSL_Pad_Size_[X]': 'F179',
                'RTP_Bump_Map_3_Probe_Mark_Inspection_USL_Pad_Size_[Y]': 'F180',
                'RTP_Bump_Map_3_Probe_Mark_Inspection_LSL_Pad_Size_[Y]': 'F181',
                'RTP_Bump_Map_3_Probe_Mark_Inspection_Pad_Mislocation_[X]': 'F182',
                'RTP_Bump_Map_3_Probe_Mark_Inspection_Pad_Mislocation_[Y]': 'F183',
                'RTP_Bump_Map_3_Probe_Mark_Inspection_Pad_Sensitivity': 'F184',
                'RTP_Bump_Map_3_Probe_Mark_Inspection_Probe__Sensitivity': 'F185',
                'RTP_Bump_Map_3_Probe_Mark_Inspection_Pad_Low_Threshold': 'F186',
                'RTP_Bump_Map_3_Probe_Mark_Inspection_Pad_High_Threshold': 'F187',
                'RTP_Bump_Map_3_Probe_Mark_Inspection_Max_Area_For_Noise_[Spots]': 'F188',
                'RTP_Bump_Map_3_Probe_Mark_Inspection_PM_Max_Area_[%_From_pad]': 'F189',
                'RTP_Bump_Map_3_Probe_Mark_Inspection_PM_Min_Area': 'F190',
                'RTP_Bump_Map_3_Probe_Mark_Inspection_Max_Number_Of_Prob_Marks': 'F191',
                'RTP_Bump_Map_3_Probe_Mark_Inspection_Min_Number_Of_Prob_Marks': 'F192',
                'RTP_Bump_Map_3_Probe_Mark_Inspection_Min_acceptable_distance__from_Pad': 'F193',
                'RTP_Bump_Map_3_Probe_Mark_Inspection_Max_PM_size_allowed_touching_the_Pad': 'F194',
                'RTP_Bump_Map_3_Probe_Mark_Inspection_Enable_surface_zone': 'F195',
                'RTP_Bump_Map_3_Probe_Mark_Inspection_Dont_Care_zone': 'F196',
                'RTP_Bump_Map_3_Probe_Mark_Inspection_Surface_Zone': 'F197',
                'RTP_Bump_Map_3_Probe_Mark_Inspection_Defect_Area_Inside_Surface': 'F198',
                'RTP_Bump_Map_3_Probe_Mark_Inspection_Contrast_Delta_-_Dark': 'F199',
                'RTP_Bump_Map_3_Probe_Mark_Inspection_Contrast_Delta_-_bright': 'F200',
                'RTP_Bump_Map_3_PMI_Advanced_USL_Pad_Size_X': 'F314',
                'RTP_Bump_Map_3_PMI_Advanced_LSL_Pad_Size_X': 'F315',
                'RTP_Bump_Map_3_PMI_Advanced_USL_Pad_Size_Y': 'F316',
                'RTP_Bump_Map_3_PMI_Advanced_LSL_Pad_Size_Y': 'F317',
                'RTP_Bump_Map_3_PMI_Advanced_Pad_Mislocation_X': 'F318',
                'RTP_Bump_Map_3_PMI_Advanced_Pad_Mislocation_Y': 'F319',
                'RTP_Bump_Map_3_PMI_Advanced_Pad_Edge_Sensitivity': 'F320',
                'RTP_Bump_Map_3_PMI_Advanced_Pad_Sensitivity': 'F321',
                'RTP_Bump_Map_3_PMI_Advanced_PM_Sensitivity': 'F322',
                'RTP_Bump_Map_3_PMI_Advanced_Pad_Gray_Level': 'F323',
                'RTP_Bump_Map_3_PMI_Advanced_Pad_Edge_Gray_Level': 'F324',
                'RTP_Bump_Map_3_PMI_Advanced_Surface_Gray_Level': 'F325',
                'RTP_Bump_Map_3_PMI_Advanced_USL_PM_Area_[%]': 'F326',
                'RTP_Bump_Map_3_PMI_Advanced_LSL_PM_Area': 'F327',
                'RTP_Bump_Map_3_PMI_Advanced_PM_Min_Spot_Area': 'F328',
                'RTP_Bump_Map_3_PMI_Advanced_Max_Number_Of_Prob_Marks': 'F329',
                'RTP_Bump_Map_3_PMI_Advanced_Min_Number_Of_Prob_Marks': 'F330',
                'RTP_Bump_Map_3_PMI_Advanced_Min_acceptable_distance__from_Pad': 'F331',
                'RTP_Bump_Map_3_PMI_Advanced_Max_PM_size_allowed_touching_the_Pad': 'F332',
                'RTP_Bump_Map_3_PMI_Advanced_Enable_surface_zone': 'F333',
                'RTP_Bump_Map_3_PMI_Advanced_Don**_Care_zone': 'F334',
                'RTP_Bump_Map_3_PMI_Advanced_Surface_Zone': 'F335',
                'RTP_Bump_Map_3_PMI_Advanced_Min_Defect_Area': 'F336',
                'RTP_Bump_Map_3_PMI_Advanced_Contrast_Delta_-_Dark': 'F337',
                'RTP_Bump_Map_3_PMI_Advanced_Contrast_Delta_-_bright': 'F335',
                'RTP_Bump_Map_3_PMI_Advanced_nspection_Sensitivity': 'F339',
                'RTP_Bump_Map_3_PMI_Advanced_Ref_Sensitivity': 'F340',

                #[Bump_Map_4]
                'RTP_Bump_Map_4_Surface_Min_Defect_Area_-_Bright': 'F76',
                'RTP_Bump_Map_4_Surface_Min_Defect_Width_-_Bright': 'F77',
                'RTP_Bump_Map_4_Surface_Min_Defect_Length_-_Bright': 'F78',
                'RTP_Bump_Map_4_Surface_Contrast_Delta_-_Bright': 'F79',
                'RTP_Bump_Map_4_Surface_Contrast_Factor_-_Bright': 'F80',
                'RTP_Bump_Map_4_Surface_Min_Defect_Area_-_Dark': 'F81',
                'RTP_Bump_Map_4_Surface_Min_Defect_Width_-_Dark': 'F82',
                'RTP_Bump_Map_4_Surface_Min_Defect_Length_-_Dark': 'F83',
                'RTP_Bump_Map_4_Surface_Contrast_Delta_-_Dark': 'F84',
                'RTP_Bump_Map_4_Surface_Contrast_Factor_-_Dark': 'F85',
                'RTP_Bump_Map_4_Surface_Big_Area_Status_-_Bright': 'F86',
                'RTP_Bump_Map_4_Surface_Big_Area_Status_-_Dark': 'F87',
                'RTP_Bump_Map_4_Surface_Cluster_Area': 'F88',
                'RTP_Bump_Map_4_Surface_Cluster_Distance': 'F89',
                'RTP_Bump_Map_4_Surface_Cluster_Diameter': 'F90',
                'RTP_Bump_Map_4_Surface_Adaptive_Histogram_Mode': 'F91',
                'RTP_Bump_Map_4_Surface_CollectForGlobalSum': 'F92',
                'RTP_Bump_Map_4_Surface_MaxAreaSum': 'F93',
                'RTP_Bump_Map_4_Surface_Zone_CD_Radius': 'F94',
                'RTP_Bump_Map_4_Surface_Dark_Zone_CD_Percent': 'F95',
                'RTP_Bump_Map_4_Surface_Bright_Zone_CD_Percent': 'F96',
                'RTP_Bump_Map_4_Surface_MaxCountSum': 'F97',
                'RTP_Bump_Map_4_Probe_Mark_Inspection_Pad_Is_Rectangle': 'F203',
                'RTP_Bump_Map_4_Probe_Mark_Inspection_USL_Pad_Size_[X]': 'F204',
                'RTP_Bump_Map_4_Probe_Mark_Inspection_LSL_Pad_Size_[X]': 'F205',
                'RTP_Bump_Map_4_Probe_Mark_Inspection_USL_Pad_Size_[Y]': 'F206',
                'RTP_Bump_Map_4_Probe_Mark_Inspection_LSL_Pad_Size_[Y]': 'F207',
                'RTP_Bump_Map_4_Probe_Mark_Inspection_Pad_Mislocation_[X]': 'F208',
                'RTP_Bump_Map_4_Probe_Mark_Inspection_Pad_Mislocation_[Y]': 'F209',
                'RTP_Bump_Map_4_Probe_Mark_Inspection_Pad_Sensitivity': 'F210',
                'RTP_Bump_Map_4_Probe_Mark_Inspection_Probe__Sensitivity': 'F211',
                'RTP_Bump_Map_4_Probe_Mark_Inspection_Pad_Low_Threshold': 'F212',
                'RTP_Bump_Map_4_Probe_Mark_Inspection_Pad_High_Threshold': 'F213',
                'RTP_Bump_Map_4_Probe_Mark_Inspection_Max_Area_For_Noise_[Spots]': 'F214',
                'RTP_Bump_Map_4_Probe_Mark_Inspection_PM_Max_Area_[%_From_pad]': 'F215',
                'RTP_Bump_Map_4_Probe_Mark_Inspection_PM_Min_Area': 'F216',
                'RTP_Bump_Map_4_Probe_Mark_Inspection_Max_Number_Of_Prob_Marks': 'F217',
                'RTP_Bump_Map_4_Probe_Mark_Inspection_Min_Number_Of_Prob_Marks': 'F218',
                'RTP_Bump_Map_4_Probe_Mark_Inspection_Min_acceptable_distance__from_Pad': 'F219',
                'RTP_Bump_Map_4_Probe_Mark_Inspection_Max_PM_size_allowed_touching_the_Pad': 'F220',
                'RTP_Bump_Map_4_Probe_Mark_Inspection_Enable_surface_zone': 'F221',
                'RTP_Bump_Map_4_Probe_Mark_Inspection_Dont_Care_zone': 'F222',
                'RTP_Bump_Map_4_Probe_Mark_Inspection_Surface_Zone': 'F223',
                'RTP_Bump_Map_4_Probe_Mark_Inspection_Defect_Area_Inside_Surface': 'F224',
                'RTP_Bump_Map_4_Probe_Mark_Inspection_Contrast_Delta_-_Dark': 'F225',
                'RTP_Bump_Map_4_Probe_Mark_Inspection_Contrast_Delta_-_bright': 'F226',
                'RTP_Bump_Map_4_PMI_Advanced_USL_Pad_Size_X': 'F343',
                'RTP_Bump_Map_4_PMI_Advanced_LSL_Pad_Size_X': 'F344',
                'RTP_Bump_Map_4_PMI_Advanced_USL_Pad_Size_Y': 'F345',
                'RTP_Bump_Map_4_PMI_Advanced_LSL_Pad_Size_Y': 'F346',
                'RTP_Bump_Map_4_PMI_Advanced_Pad_Mislocation_X': 'F347',
                'RTP_Bump_Map_4_PMI_Advanced_Pad_Mislocation_Y': 'F348',
                'RTP_Bump_Map_4_PMI_Advanced_Pad_Edge_Sensitivity': 'F349',
                'RTP_Bump_Map_4_PMI_Advanced_Pad_Sensitivity': 'F350',
                'RTP_Bump_Map_4_PMI_Advanced_PM_Sensitivity': 'F351',
                'RTP_Bump_Map_4_PMI_Advanced_Pad_Gray_Level': 'F351',
                'RTP_Bump_Map_4_PMI_Advanced_Pad_Edge_Gray_Level': 'F353',
                'RTP_Bump_Map_4_PMI_Advanced_Surface_Gray_Level': 'F354',
                'RTP_Bump_Map_4_PMI_Advanced_USL_PM_Area_[%]': 'F355',
                'RTP_Bump_Map_4_PMI_Advanced_LSL_PM_Area': 'F356',
                'RTP_Bump_Map_4_PMI_Advanced_PM_Min_Spot_Area': 'F357',
                'RTP_Bump_Map_4_PMI_Advanced_Max_Number_Of_Prob_Marks': 'F358',
                'RTP_Bump_Map_4_PMI_Advanced_Min_Number_Of_Prob_Marks': 'F359',
                'RTP_Bump_Map_4_PMI_Advanced_Min_acceptable_distance__from_Pad': 'F360',
                'RTP_Bump_Map_4_PMI_Advanced_Max_PM_size_allowed_touching_the_Pad': 'F361',
                'RTP_Bump_Map_4_PMI_Advanced_Enable_surface_zone': 'F362',
                'RTP_Bump_Map_4_PMI_Advanced_Don**_Care_zone': 'F363',
                'RTP_Bump_Map_4_PMI_Advanced_Surface_Zone': 'F364',
                'RTP_Bump_Map_4_PMI_Advanced_Min_Defect_Area': 'F365',
                'RTP_Bump_Map_4_PMI_Advanced_Contrast_Delta_-_Dark': 'F366',
                'RTP_Bump_Map_4_PMI_Advanced_Contrast_Delta_-_bright': 'F367',
                'RTP_Bump_Map_4_PMI_Advanced_nspection_Sensitivity': 'F368',
                'RTP_Bump_Map_4_PMI_Advanced_Ref_Sensitivity': 'F369',

                #[Bump_Map_5]
                'RTP_Bump_Map_5_Surface_Min_Defect_Area_-_Bright': 'F100',
                'RTP_Bump_Map_5_Surface_Min_Defect_Width_-_Bright': 'F101',
                'RTP_Bump_Map_5_Surface_Min_Defect_Length_-_Bright': 'F102',
                'RTP_Bump_Map_5_Surface_Contrast_Delta_-_Bright': 'F103',
                'RTP_Bump_Map_5_Surface_Contrast_Factor_-_Bright': 'F104',
                'RTP_Bump_Map_5_Surface_Min_Defect_Area_-_Dark': 'F105',
                'RTP_Bump_Map_5_Surface_Min_Defect_Width_-_Dark': 'F106',
                'RTP_Bump_Map_5_Surface_Min_Defect_Length_-_Dark': 'F107',
                'RTP_Bump_Map_5_Surface_Contrast_Delta_-_Dark': 'F108',
                'RTP_Bump_Map_5_Surface_Contrast_Factor_-_Dark': 'F109',
                'RTP_Bump_Map_5_Surface_Big_Area_Status_-_Bright': 'F110',
                'RTP_Bump_Map_5_Surface_Big_Area_Status_-_Dark': 'F111',
                'RTP_Bump_Map_5_Surface_Cluster_Area': 'F112',
                'RTP_Bump_Map_5_Surface_Cluster_Distance': 'F113',
                'RTP_Bump_Map_5_Surface_Cluster_Diameter': 'F114',
                'RTP_Bump_Map_5_Surface_Adaptive_Histogram_Mode': 'F115',
                'RTP_Bump_Map_5_Surface_CollectForGlobalSum': 'F116',
                'RTP_Bump_Map_5_Surface_MaxAreaSum': 'F117',
                'RTP_Bump_Map_5_Surface_Zone_CD_Radius': 'F118',
                'RTP_Bump_Map_5_Surface_Dark_Zone_CD_Percent': 'F119',
                'RTP_Bump_Map_5_Surface_Bright_Zone_CD_Percent': 'F120',
                'RTP_Bump_Map_5_Surface_MaxCountSum': 'F121',
                'RTP_Bump_Map_5_Probe_Mark_Inspection_Pad_Is_Rectangle': 'F229',
                'RTP_Bump_Map_5_Probe_Mark_Inspection_USL_Pad_Size_[X]': 'F230',
                'RTP_Bump_Map_5_Probe_Mark_Inspection_LSL_Pad_Size_[X]': 'F231',
                'RTP_Bump_Map_5_Probe_Mark_Inspection_USL_Pad_Size_[Y]': 'F232',
                'RTP_Bump_Map_5_Probe_Mark_Inspection_LSL_Pad_Size_[Y]': 'F233',
                'RTP_Bump_Map_5_Probe_Mark_Inspection_Pad_Mislocation_[X]': 'F234',
                'RTP_Bump_Map_5_Probe_Mark_Inspection_Pad_Mislocation_[Y]': 'F235',
                'RTP_Bump_Map_5_Probe_Mark_Inspection_Pad_Sensitivity': 'F236',
                'RTP_Bump_Map_5_Probe_Mark_Inspection_Probe__Sensitivity': 'F237',
                'RTP_Bump_Map_5_Probe_Mark_Inspection_Pad_Low_Threshold': 'F238',
                'RTP_Bump_Map_5_Probe_Mark_Inspection_Pad_High_Threshold': 'F239',
                'RTP_Bump_Map_5_Probe_Mark_Inspection_Max_Area_For_Noise_[Spots]': 'F240',
                'RTP_Bump_Map_5_Probe_Mark_Inspection_PM_Max_Area_[%_From_pad]': 'F241',
                'RTP_Bump_Map_5_Probe_Mark_Inspection_PM_Min_Area': 'F242',
                'RTP_Bump_Map_5_Probe_Mark_Inspection_Max_Number_Of_Prob_Marks': 'F243',
                'RTP_Bump_Map_5_Probe_Mark_Inspection_Min_Number_Of_Prob_Marks': 'F244',
                'RTP_Bump_Map_5_Probe_Mark_Inspection_Min_acceptable_distance__from_Pad': 'F245',
                'RTP_Bump_Map_5_Probe_Mark_Inspection_Max_PM_size_allowed_touching_the_Pad': 'F246',
                'RTP_Bump_Map_5_Probe_Mark_Inspection_Enable_surface_zone': 'F247',
                'RTP_Bump_Map_5_Probe_Mark_Inspection_Dont_Care_zone': 'F248',
                'RTP_Bump_Map_5_Probe_Mark_Inspection_Surface_Zone': 'F249',
                'RTP_Bump_Map_5_Probe_Mark_Inspection_Defect_Area_Inside_Surface': 'F250',
                'RTP_Bump_Map_5_Probe_Mark_Inspection_Contrast_Delta_-_Dark': 'F251',
                'RTP_Bump_Map_5_Probe_Mark_Inspection_Contrast_Delta_-_bright': 'F252',
                'RTP_Bump_Map_5_PMI_Advanced_USL_Pad_Size_X': 'F372',
                'RTP_Bump_Map_5_PMI_Advanced_LSL_Pad_Size_X': 'F373',
                'RTP_Bump_Map_5_PMI_Advanced_USL_Pad_Size_Y': 'F374',
                'RTP_Bump_Map_5_PMI_Advanced_LSL_Pad_Size_Y': 'F375',
                'RTP_Bump_Map_5_PMI_Advanced_Pad_Mislocation_X': 'F376',
                'RTP_Bump_Map_5_PMI_Advanced_Pad_Mislocation_Y': 'F377',
                'RTP_Bump_Map_5_PMI_Advanced_Pad_Edge_Sensitivity': 'F378',
                'RTP_Bump_Map_5_PMI_Advanced_Pad_Sensitivity': 'F379',
                'RTP_Bump_Map_5_PMI_Advanced_PM_Sensitivity': 'F380',
                'RTP_Bump_Map_5_PMI_Advanced_Pad_Gray_Level': 'F381',
                'RTP_Bump_Map_5_PMI_Advanced_Pad_Edge_Gray_Level': 'F382',
                'RTP_Bump_Map_5_PMI_Advanced_Surface_Gray_Level': 'F383',
                'RTP_Bump_Map_5_PMI_Advanced_USL_PM_Area_[%]': 'F384',
                'RTP_Bump_Map_5_PMI_Advanced_LSL_PM_Area': 'F385',
                'RTP_Bump_Map_5_PMI_Advanced_PM_Min_Spot_Area': 'F386',
                'RTP_Bump_Map_5_PMI_Advanced_Max_Number_Of_Prob_Marks': 'F387',
                'RTP_Bump_Map_5_PMI_Advanced_Min_Number_Of_Prob_Marks': 'F388',
                'RTP_Bump_Map_5_PMI_Advanced_Min_acceptable_distance__from_Pad': 'F389',
                'RTP_Bump_Map_5_PMI_Advanced_Max_PM_size_allowed_touching_the_Pad': 'F390',
                'RTP_Bump_Map_5_PMI_Advanced_Enable_surface_zone': 'F391',
                'RTP_Bump_Map_5_PMI_Advanced_Don**_Care_zone': 'F392',
                'RTP_Bump_Map_5_PMI_Advanced_Surface_Zone': 'F393',
                'RTP_Bump_Map_5_PMI_Advanced_Min_Defect_Area': 'F394',
                'RTP_Bump_Map_5_PMI_Advanced_Contrast_Delta_-_Dark': 'F395',
                'RTP_Bump_Map_5_PMI_Advanced_Contrast_Delta_-_bright': 'F396',
                'RTP_Bump_Map_5_PMI_Advanced_nspection_Sensitivity': 'F397',
                'RTP_Bump_Map_5_PMI_Advanced_Ref_Sensitivity': 'F398'
            }

            bump_device_mappings = {
                #[Bump_Map_1]
                'RTP_Bump_Map_1_Solder_Bump_Bump_Color_is_White': 'F4',
                'RTP_Bump_Map_1_Solder_Bump_Bump_is_Contaminated': 'F5',
                'RTP_Bump_Map_1_Solder_Bump_Bump_Diamter_LSL': 'F6',
                'RTP_Bump_Map_1_Solder_Bump_Bump_Diamter_USL': 'F7',
                'RTP_Bump_Map_1_Solder_Bump_Mislocation_X': 'F8',
                'RTP_Bump_Map_1_Solder_Bump_Mislocation_Y': 'F9',
                'RTP_Bump_Map_1_Solder_Bump_Detection_Threshold': 'F10',
                'RTP_Bump_Map_1_Solder_Bump_Detection_Gradient': 'F11',
                'RTP_Bump_Map_1_Solder_Bump_Bump_Roundness': 'F12',
                'RTP_Bump_Map_1_Solder_Bump_Number_Of_Lines': 'F13',
                'RTP_Bump_Map_1_Solder_Bump_Min_Points_for_bump_detection': 'F14',
                'RTP_Bump_Map_1_Solder_Bump_RadiusPercentIn': 'F15',
                'RTP_Bump_Map_1_Solder_Bump_RadiusPercentOut': 'F16',
                'RTP_Bump_Map_1_Solder_Bump_LSL_ShapeViolation': 'F17',
                'RTP_Bump_Map_1_Solder_Bump_USL_ShapeViolation': 'F18',
                'RTP_Bump_Map_1_Solder_Bump_EdgeDetectThreshold': 'F19',
                'RTP_Bump_Map_1_Solder_Bump_EdgeDetectArea': 'F20',
                'RTP_Bump_Map_1_Solder_Bump_EdgeDetectLength': 'F21',
                'RTP_Bump_Map_1_Solder_Bump_EdgeDetectDiameter': 'F22',
                'RTP_Bump_Map_1_Solder_Bump_Edge_-_MinGL': 'F23',
                'RTP_Bump_Map_1_Solder_Bump_Edge_-_MaxGL': 'F24',
                'RTP_Bump_Map_1_Solder_Bump_Mislocation': 'F25',
                'RTP_Bump_Map_1_Surface_on_SB_Enable_Surface_Moving': 'F27',
                'RTP_Bump_Map_1_Surface_on_SB_Exposed_Area_High_TH': 'F28',
                'RTP_Bump_Map_1_Surface_on_SB_Exposed_Area_Low_TH': 'F29',
                'RTP_Bump_Map_1_Surface_on_SB_Actual__position_don\'t_care_width': 'F30',
                'RTP_Bump_Map_1_Surface_on_SB_Original_position_don\'t_care_width': 'F31',
                'RTP_Bump_Map_1_Surface_on_SB_Min_Defect_Area_-_Bright': 'F32',
                'RTP_Bump_Map_1_Surface_on_SB_Min_Defect_Width_-_Bright': 'F33',
                'RTP_Bump_Map_1_Surface_on_SB_Min_Defect_Length_-_Bright': 'F34',
                'RTP_Bump_Map_1_Surface_on_SB_Contrast_Delta_-_Bright': 'F35',
                'RTP_Bump_Map_1_Surface_on_SB_Min_Defect_Area_-_Dark': 'F36',
                'RTP_Bump_Map_1_Surface_on_SB_Min_Defect_Width_-_Dark': 'F37',
                'RTP_Bump_Map_1_Surface_on_SB_Min_Defect_Length_-_Dark': 'F38',
                'RTP_Bump_Map_1_Surface_on_SB_Contrast_Delta_-_Dark': 'F39',
                'RTP_Bump_Map_1_Surface_on_SB_Elongation': 'F40',
                'RTP_Bump_Map_1_Surface_on_SB_MaxAreaSum': 'F41',
                'RTP_Bump_Map_1_Surface_on_SB_CollectForGlobalSum': 'F42',
                'RTP_Bump_Map_1_Uniform_Surface_on_SB_Enable_Moving_Surface': 'F44',
                'RTP_Bump_Map_1_Uniform_Surface_on_SB_Exposed_Area_High_TH': 'F45',
                'RTP_Bump_Map_1_Uniform_Surface_on_SB_Exposed_Area_Low_TH': 'F46',
                'RTP_Bump_Map_1_Uniform_Surface_on_SB_Position_Don\'t-Care_Width': 'F47',
                'RTP_Bump_Map_1_Uniform_Surface_on_SB_Original_position_don\'t_care_width': 'F48',
                'RTP_Bump_Map_1_Uniform_Surface_on_SB_Min_Defect_Area_-_Bright': 'F49',
                'RTP_Bump_Map_1_Uniform_Surface_on_SB_Min_Defect_Width_-_Bright': 'F50',
                'RTP_Bump_Map_1_Uniform_Surface_on_SB_Min_Defect_Length_-_Bright': 'F51',
                'RTP_Bump_Map_1_Uniform_Surface_on_SB_Contrast_Upper_value_-_Bright': 'F52',
                'RTP_Bump_Map_1_Uniform_Surface_on_SB_Min_Defect_Area_-_Dark': 'F53',
                'RTP_Bump_Map_1_Uniform_Surface_on_SB_Min_Defect_Width_-_Dark': 'F54',
                'RTP_Bump_Map_1_Uniform_Surface_on_SB_Min_Defect_Length_-_Dark': 'F55',
                'RTP_Bump_Map_1_Uniform_Surface_on_SB_Contrast_Lower_value_-_Dark': 'F56',
                'RTP_Bump_Map_1_Uniform_Surface_on_SB_MaxAreaSum': 'F57',
                'RTP_Bump_Map_1_Uniform_Surface_on_SB_CollectForGlobalSum': 'F58',

                #[Bump_Map_2]
                'RTP_Bump_Map_2_Solder_Bump_Bump_Color_is_White': 'F62',
                'RTP_Bump_Map_2_Solder_Bump_Bump_is_Contaminated': 'F63',
                'RTP_Bump_Map_2_Solder_Bump_Bump_Diamter_LSL': 'F64',
                'RTP_Bump_Map_2_Solder_Bump_Bump_Diamter_USL': 'F65',
                'RTP_Bump_Map_2_Solder_Bump_Mislocation_X': 'F66',
                'RTP_Bump_Map_2_Solder_Bump_Mislocation_Y': 'F67',
                'RTP_Bump_Map_2_Solder_Bump_Detection_Threshold': 'F68',
                'RTP_Bump_Map_2_Solder_Bump_Detection_Gradient': 'F69',
                'RTP_Bump_Map_2_Solder_Bump_Bump_Roundness': 'F70',
                'RTP_Bump_Map_2_Solder_Bump_Number_Of_Lines': 'F71',
                'RTP_Bump_Map_2_Solder_Bump_Min_Points_for_bump_detection': 'F72',
                'RTP_Bump_Map_2_Solder_Bump_RadiusPercentIn': 'F73',
                'RTP_Bump_Map_2_Solder_Bump_RadiusPercentOut': 'F74',
                'RTP_Bump_Map_2_Solder_Bump_LSL_ShapeViolation': 'F75',
                'RTP_Bump_Map_2_Solder_Bump_USL_ShapeViolation': 'F76',
                'RTP_Bump_Map_2_Solder_Bump_EdgeDetectThreshold': 'F77',
                'RTP_Bump_Map_2_Solder_Bump_EdgeDetectArea': 'F78',
                'RTP_Bump_Map_2_Solder_Bump_EdgeDetectLength': 'F79',
                'RTP_Bump_Map_2_Solder_Bump_EdgeDetectDiameter': 'F80',
                'RTP_Bump_Map_2_Solder_Bump_Edge_-_MinGL': 'F81',
                'RTP_Bump_Map_2_Solder_Bump_Edge_-_MaxGL': 'F82',
                'RTP_Bump_Map_2_Solder_Bump_Mislocation': 'F83',
                'RTP_Bump_Map_2_Surface_on_SB_Enable_Surface_Moving': 'F85',
                'RTP_Bump_Map_2_Surface_on_SB_Exposed_Area_High_TH': 'F86',
                'RTP_Bump_Map_2_Surface_on_SB_Exposed_Area_Low_TH': 'F87',
                'RTP_Bump_Map_2_Surface_on_SB_Actual__position_don\'t_care_width': 'F88',
                'RTP_Bump_Map_2_Surface_on_SB_Original_position_don\'t_care_width': 'F89',
                'RTP_Bump_Map_2_Surface_on_SB_Min_Defect_Area_-_Bright': 'F90',
                'RTP_Bump_Map_2_Surface_on_SB_Min_Defect_Width_-_Bright': 'F91',
                'RTP_Bump_Map_2_Surface_on_SB_Min_Defect_Length_-_Bright': 'F92',
                'RTP_Bump_Map_2_Surface_on_SB_Contrast_Delta_-_Bright': 'F93',
                'RTP_Bump_Map_2_Surface_on_SB_Min_Defect_Area_-_Dark': 'F94',
                'RTP_Bump_Map_2_Surface_on_SB_Min_Defect_Width_-_Dark': 'F95',
                'RTP_Bump_Map_2_Surface_on_SB_Min_Defect_Length_-_Dark': 'F96',
                'RTP_Bump_Map_2_Surface_on_SB_Contrast_Delta_-_Dark': 'F97',
                'RTP_Bump_Map_2_Surface_on_SB_Elongation': 'F98',
                'RTP_Bump_Map_2_Surface_on_SB_MaxAreaSum': 'F99',
                'RTP_Bump_Map_2_Surface_on_SB_CollectForGlobalSum': 'F100',
                'RTP_Bump_Map_2_Uniform_Surface_on_SB_Enable_Moving_Surface': 'F102',
                'RTP_Bump_Map_2_Uniform_Surface_on_SB_Exposed_Area_High_TH': 'F103',
                'RTP_Bump_Map_2_Uniform_Surface_on_SB_Exposed_Area_Low_TH': 'F104',
                'RTP_Bump_Map_2_Uniform_Surface_on_SB_Position_Don\'t-Care_Width': 'F105',
                'RTP_Bump_Map_2_Uniform_Surface_on_SB_Original_position_don\'t_care_width': 'F106',
                'RTP_Bump_Map_2_Uniform_Surface_on_SB_Min_Defect_Area_-_Bright': 'F107',
                'RTP_Bump_Map_2_Uniform_Surface_on_SB_Min_Defect_Width_-_Bright': 'F108',
                'RTP_Bump_Map_2_Uniform_Surface_on_SB_Min_Defect_Length_-_Bright': 'F109',
                'RTP_Bump_Map_2_Uniform_Surface_on_SB_Contrast_Upper_value_-_Bright': 'F110',
                'RTP_Bump_Map_2_Uniform_Surface_on_SB_Min_Defect_Area_-_Dark': 'F111',
                'RTP_Bump_Map_2_Uniform_Surface_on_SB_Min_Defect_Width_-_Dark': 'F112',
                'RTP_Bump_Map_2_Uniform_Surface_on_SB_Min_Defect_Length_-_Dark': 'F113',
                'RTP_Bump_Map_2_Uniform_Surface_on_SB_Contrast_Lower_value_-_Dark': 'F114',
                'RTP_Bump_Map_2_Uniform_Surface_on_SB_MaxAreaSum': 'F115',
                'RTP_Bump_Map_2_Uniform_Surface_on_SB_CollectForGlobalSum': 'F116',

                #[Bump_Map_3]
                'RTP_Bump_Map_3_Solder_Bump_Bump_Color_is_White': 'F120',
                'RTP_Bump_Map_3_Solder_Bump_Bump_is_Contaminated': 'F121',
                'RTP_Bump_Map_3_Solder_Bump_Bump_Diamter_LSL': 'F122',
                'RTP_Bump_Map_3_Solder_Bump_Bump_Diamter_USL': 'F123',
                'RTP_Bump_Map_3_Solder_Bump_Mislocation_X': 'F124',
                'RTP_Bump_Map_3_Solder_Bump_Mislocation_Y': 'F125',
                'RTP_Bump_Map_3_Solder_Bump_Detection_Threshold': 'F126',
                'RTP_Bump_Map_3_Solder_Bump_Detection_Gradient': 'F127',
                'RTP_Bump_Map_3_Solder_Bump_Bump_Roundness': 'F128',
                'RTP_Bump_Map_3_Solder_Bump_Number_Of_Lines': 'F129',
                'RTP_Bump_Map_3_Solder_Bump_Min_Points_for_bump_detection': 'F130',
                'RTP_Bump_Map_3_Solder_Bump_RadiusPercentIn': 'F131',
                'RTP_Bump_Map_3_Solder_Bump_RadiusPercentOut': 'F132',
                'RTP_Bump_Map_3_Solder_Bump_LSL_ShapeViolation': 'F133',
                'RTP_Bump_Map_3_Solder_Bump_USL_ShapeViolation': 'F134',
                'RTP_Bump_Map_3_Solder_Bump_EdgeDetectThreshold': 'F135',
                'RTP_Bump_Map_3_Solder_Bump_EdgeDetectArea': 'F136',
                'RTP_Bump_Map_3_Solder_Bump_EdgeDetectLength': 'F137',
                'RTP_Bump_Map_3_Solder_Bump_EdgeDetectDiameter': 'F138',
                'RTP_Bump_Map_3_Solder_Bump_Edge_-_MinGL': 'F139',
                'RTP_Bump_Map_3_Solder_Bump_Edge_-_MaxGL': 'F140',
                'RTP_Bump_Map_3_Solder_Bump_Mislocation': 'F141',
                'RTP_Bump_Map_3_Surface_on_SB_Enable_Surface_Moving': 'F143',
                'RTP_Bump_Map_3_Surface_on_SB_Exposed_Area_High_TH': 'F144',
                'RTP_Bump_Map_3_Surface_on_SB_Exposed_Area_Low_TH': 'F145',
                'RTP_Bump_Map_3_Surface_on_SB_Actual__position_don\'t_care_width': 'F146',
                'RTP_Bump_Map_3_Surface_on_SB_Original_position_don\'t_care_width': 'F147',
                'RTP_Bump_Map_3_Surface_on_SB_Min_Defect_Area_-_Bright': 'F148',
                'RTP_Bump_Map_3_Surface_on_SB_Min_Defect_Width_-_Bright': 'F149',
                'RTP_Bump_Map_3_Surface_on_SB_Min_Defect_Length_-_Bright': 'F150',
                'RTP_Bump_Map_3_Surface_on_SB_Contrast_Delta_-_Bright': 'F151',
                'RTP_Bump_Map_3_Surface_on_SB_Min_Defect_Area_-_Dark': 'F152',
                'RTP_Bump_Map_3_Surface_on_SB_Min_Defect_Width_-_Dark': 'F153',
                'RTP_Bump_Map_3_Surface_on_SB_Min_Defect_Length_-_Dark': 'F154',
                'RTP_Bump_Map_3_Surface_on_SB_Contrast_Delta_-_Dark': 'F155',
                'RTP_Bump_Map_3_Surface_on_SB_Elongation': 'F156',
                'RTP_Bump_Map_3_Surface_on_SB_MaxAreaSum': 'F157',
                'RTP_Bump_Map_3_Surface_on_SB_CollectForGlobalSum': 'F158',
                'RTP_Bump_Map_3_Uniform_Surface_on_SB_Enable_Moving_Surface': 'F160',
                'RTP_Bump_Map_3_Uniform_Surface_on_SB_Exposed_Area_High_TH': 'F161',
                'RTP_Bump_Map_3_Uniform_Surface_on_SB_Exposed_Area_Low_TH': 'F162',
                'RTP_Bump_Map_3_Uniform_Surface_on_SB_Position_Don\'t-Care_Width': 'F163',
                'RTP_Bump_Map_3_Uniform_Surface_on_SB_Original_position_don\'t_care_width': 'F164',
                'RTP_Bump_Map_3_Uniform_Surface_on_SB_Min_Defect_Area_-_Bright': 'F165',
                'RTP_Bump_Map_3_Uniform_Surface_on_SB_Min_Defect_Width_-_Bright': 'F166',
                'RTP_Bump_Map_3_Uniform_Surface_on_SB_Min_Defect_Length_-_Bright': 'F167',
                'RTP_Bump_Map_3_Uniform_Surface_on_SB_Contrast_Upper_value_-_Bright': 'F168',
                'RTP_Bump_Map_3_Uniform_Surface_on_SB_Min_Defect_Area_-_Dark': 'F169',
                'RTP_Bump_Map_3_Uniform_Surface_on_SB_Min_Defect_Width_-_Dark': 'F170',
                'RTP_Bump_Map_3_Uniform_Surface_on_SB_Min_Defect_Length_-_Dark': 'F171',
                'RTP_Bump_Map_3_Uniform_Surface_on_SB_Contrast_Lower_value_-_Dark': 'F172',
                'RTP_Bump_Map_3_Uniform_Surface_on_SB_MaxAreaSum': 'F173',
                'RTP_Bump_Map_3_Uniform_Surface_on_SB_CollectForGlobalSum': 'F174',

                #[Bump_Map_4]
                'RTP_Bump_Map_4_Solder_Bump_Bump_Color_is_White': 'F178',
                'RTP_Bump_Map_4_Solder_Bump_Bump_is_Contaminated': 'F179',
                'RTP_Bump_Map_4_Solder_Bump_Bump_Diamter_LSL': 'F180',
                'RTP_Bump_Map_4_Solder_Bump_Bump_Diamter_USL': 'F181',
                'RTP_Bump_Map_4_Solder_Bump_Mislocation_X': 'F182',
                'RTP_Bump_Map_4_Solder_Bump_Mislocation_Y': 'F183',
                'RTP_Bump_Map_4_Solder_Bump_Detection_Threshold': 'F184',
                'RTP_Bump_Map_4_Solder_Bump_Detection_Gradient': 'F185',
                'RTP_Bump_Map_4_Solder_Bump_Bump_Roundness': 'F186',
                'RTP_Bump_Map_4_Solder_Bump_Number_Of_Lines': 'F187',
                'RTP_Bump_Map_4_Solder_Bump_Min_Points_for_bump_detection': 'F188',
                'RTP_Bump_Map_4_Solder_Bump_RadiusPercentIn': 'F189',
                'RTP_Bump_Map_4_Solder_Bump_RadiusPercentOut': 'F190',
                'RTP_Bump_Map_4_Solder_Bump_LSL_ShapeViolation': 'F191',
                'RTP_Bump_Map_4_Solder_Bump_USL_ShapeViolation': 'F192',
                'RTP_Bump_Map_4_Solder_Bump_EdgeDetectThreshold': 'F193',
                'RTP_Bump_Map_4_Solder_Bump_EdgeDetectArea': 'F194',
                'RTP_Bump_Map_4_Solder_Bump_EdgeDetectLength': 'F195',
                'RTP_Bump_Map_4_Solder_Bump_EdgeDetectDiameter': 'F196',
                'RTP_Bump_Map_4_Solder_Bump_Edge_-_MinGL': 'F197',
                'RTP_Bump_Map_4_Solder_Bump_Edge_-_MaxGL': 'F198',
                'RTP_Bump_Map_4_Solder_Bump_Mislocation': 'F199',
                'RTP_Bump_Map_4_Surface_on_SB_Enable_Surface_Moving': 'F201',
                'RTP_Bump_Map_4_Surface_on_SB_Exposed_Area_High_TH': 'F202',
                'RTP_Bump_Map_4_Surface_on_SB_Exposed_Area_Low_TH': 'F203',
                'RTP_Bump_Map_4_Surface_on_SB_Actual__position_don\'t_care_width': 'F204',
                'RTP_Bump_Map_4_Surface_on_SB_Original_position_don\'t_care_width': 'F205',
                'RTP_Bump_Map_4_Surface_on_SB_Min_Defect_Area_-_Bright': 'F206',
                'RTP_Bump_Map_4_Surface_on_SB_Min_Defect_Width_-_Bright': 'F207',
                'RTP_Bump_Map_4_Surface_on_SB_Min_Defect_Length_-_Bright': 'F208',
                'RTP_Bump_Map_4_Surface_on_SB_Contrast_Delta_-_Bright': 'F209',
                'RTP_Bump_Map_4_Surface_on_SB_Min_Defect_Area_-_Dark': 'F210',
                'RTP_Bump_Map_4_Surface_on_SB_Min_Defect_Width_-_Dark': 'F211',
                'RTP_Bump_Map_4_Surface_on_SB_Min_Defect_Length_-_Dark': 'F212',
                'RTP_Bump_Map_4_Surface_on_SB_Contrast_Delta_-_Dark': 'F213',
                'RTP_Bump_Map_4_Surface_on_SB_Elongation': 'F214',
                'RTP_Bump_Map_4_Surface_on_SB_MaxAreaSum': 'F215',
                'RTP_Bump_Map_4_Surface_on_SB_CollectForGlobalSum': 'F216',
                'RTP_Bump_Map_4_Uniform_Surface_on_SB_Enable_Moving_Surface': 'F218',
                'RTP_Bump_Map_4_Uniform_Surface_on_SB_Exposed_Area_High_TH': 'F219',
                'RTP_Bump_Map_4_Uniform_Surface_on_SB_Exposed_Area_Low_TH': 'F220',
                'RTP_Bump_Map_4_Uniform_Surface_on_SB_Position_Don\'t-Care_Width': 'F221',
                'RTP_Bump_Map_4_Uniform_Surface_on_SB_Original_position_don\'t_care_width': 'F222',
                'RTP_Bump_Map_4_Uniform_Surface_on_SB_Min_Defect_Area_-_Bright': 'F223',
                'RTP_Bump_Map_4_Uniform_Surface_on_SB_Min_Defect_Width_-_Bright': 'F224',
                'RTP_Bump_Map_4_Uniform_Surface_on_SB_Min_Defect_Length_-_Bright': 'F225',
                'RTP_Bump_Map_4_Uniform_Surface_on_SB_Contrast_Upper_value_-_Bright': 'F226',
                'RTP_Bump_Map_4_Uniform_Surface_on_SB_Min_Defect_Area_-_Dark': 'F227',
                'RTP_Bump_Map_4_Uniform_Surface_on_SB_Min_Defect_Width_-_Dark': 'F228',
                'RTP_Bump_Map_4_Uniform_Surface_on_SB_Min_Defect_Length_-_Dark': 'F229',
                'RTP_Bump_Map_4_Uniform_Surface_on_SB_Contrast_Lower_value_-_Dark': 'F230',
                'RTP_Bump_Map_4_Uniform_Surface_on_SB_MaxAreaSum': 'F231',
                'RTP_Bump_Map_4_Uniform_Surface_on_SB_CollectForGlobalSum': 'F232',

                #[Bump_Map_5]
                'RTP_Bump_Map_5_Solder_Bump_Bump_Color_is_White': 'F236',
                'RTP_Bump_Map_5_Solder_Bump_Bump_is_Contaminated': 'F237',
                'RTP_Bump_Map_5_Solder_Bump_Bump_Diamter_LSL': 'F238',
                'RTP_Bump_Map_5_Solder_Bump_Bump_Diamter_USL': 'F239',
                'RTP_Bump_Map_5_Solder_Bump_Mislocation_X': 'F240',
                'RTP_Bump_Map_5_Solder_Bump_Mislocation_Y': 'F241',
                'RTP_Bump_Map_5_Solder_Bump_Detection_Threshold': 'F242',
                'RTP_Bump_Map_5_Solder_Bump_Detection_Gradient': 'F243',
                'RTP_Bump_Map_5_Solder_Bump_Bump_Roundness': 'F244',
                'RTP_Bump_Map_5_Solder_Bump_Number_Of_Lines': 'F245',
                'RTP_Bump_Map_5_Solder_Bump_Min_Points_for_bump_detection': 'F246',
                'RTP_Bump_Map_5_Solder_Bump_RadiusPercentIn': 'F247',
                'RTP_Bump_Map_5_Solder_Bump_RadiusPercentOut': 'F248',
                'RTP_Bump_Map_5_Solder_Bump_LSL_ShapeViolation': 'F249',
                'RTP_Bump_Map_5_Solder_Bump_USL_ShapeViolation': 'F250',
                'RTP_Bump_Map_5_Solder_Bump_EdgeDetectThreshold': 'F251',
                'RTP_Bump_Map_5_Solder_Bump_EdgeDetectArea': 'F252',
                'RTP_Bump_Map_5_Solder_Bump_EdgeDetectLength': 'F253',
                'RTP_Bump_Map_5_Solder_Bump_EdgeDetectDiameter': 'F254',
                'RTP_Bump_Map_5_Solder_Bump_Edge_-_MinGL': 'F255',
                'RTP_Bump_Map_5_Solder_Bump_Edge_-_MaxGL': 'F256',
                'RTP_Bump_Map_5_Solder_Bump_Mislocation': 'F257',
                'RTP_Bump_Map_5_Surface_on_SB_Enable_Surface_Moving': 'F259',
                'RTP_Bump_Map_5_Surface_on_SB_Exposed_Area_High_TH': 'F260',
                'RTP_Bump_Map_5_Surface_on_SB_Exposed_Area_Low_TH': 'F261',
                'RTP_Bump_Map_5_Surface_on_SB_Actual__position_don\'t_care_width': 'F262',
                'RTP_Bump_Map_5_Surface_on_SB_Original_position_don\'t_care_width': 'F263',
                'RTP_Bump_Map_5_Surface_on_SB_Min_Defect_Area_-_Bright': 'F264',
                'RTP_Bump_Map_5_Surface_on_SB_Min_Defect_Width_-_Bright': 'F265',
                'RTP_Bump_Map_5_Surface_on_SB_Min_Defect_Length_-_Bright': 'F266',
                'RTP_Bump_Map_5_Surface_on_SB_Contrast_Delta_-_Bright': 'F267',
                'RTP_Bump_Map_5_Surface_on_SB_Min_Defect_Area_-_Dark': 'F268',
                'RTP_Bump_Map_5_Surface_on_SB_Min_Defect_Width_-_Dark': 'F269',
                'RTP_Bump_Map_5_Surface_on_SB_Min_Defect_Length_-_Dark': 'F270',
                'RTP_Bump_Map_5_Surface_on_SB_Contrast_Delta_-_Dark': 'F271',
                'RTP_Bump_Map_5_Surface_on_SB_Elongation': 'F272',
                'RTP_Bump_Map_5_Surface_on_SB_MaxAreaSum': 'F273',
                'RTP_Bump_Map_5_Surface_on_SB_CollectForGlobalSum': 'F274',
                'RTP_Bump_Map_5_Uniform_Surface_on_SB_Enable_Moving_Surface': 'F276',
                'RTP_Bump_Map_5_Uniform_Surface_on_SB_Exposed_Area_High_TH': 'F277',
                'RTP_Bump_Map_5_Uniform_Surface_on_SB_Exposed_Area_Low_TH': 'F278',
                'RTP_Bump_Map_5_Uniform_Surface_on_SB_Position_Don\'t-Care_Width': 'F279',
                'RTP_Bump_Map_5_Uniform_Surface_on_SB_Original_position_don\'t_care_width': 'F280',
                'RTP_Bump_Map_5_Uniform_Surface_on_SB_Min_Defect_Area_-_Bright': 'F281',
                'RTP_Bump_Map_5_Uniform_Surface_on_SB_Min_Defect_Width_-_Bright': 'F282',
                'RTP_Bump_Map_5_Uniform_Surface_on_SB_Min_Defect_Length_-_Bright': 'F283',
                'RTP_Bump_Map_5_Uniform_Surface_on_SB_Contrast_Upper_value_-_Bright': 'F284',
                'RTP_Bump_Map_5_Uniform_Surface_on_SB_Min_Defect_Area_-_Dark': 'F285',
                'RTP_Bump_Map_5_Uniform_Surface_on_SB_Min_Defect_Width_-_Dark': 'F286',
                'RTP_Bump_Map_5_Uniform_Surface_on_SB_Min_Defect_Length_-_Dark': 'F287',
                'RTP_Bump_Map_5_Uniform_Surface_on_SB_Contrast_Lower_value_-_Dark': 'F288',
                'RTP_Bump_Map_5_Uniform_Surface_on_SB_MaxAreaSum': 'F289',
                'RTP_Bump_Map_5_Uniform_Surface_on_SB_CollectForGlobalSum': 'F290',
            }

            all_mappings = {
                'Default': {
                    "Check list": check_list_mappings,
                    "Surface": surface_mappings,
                    "Pad device": pad_device_mappings,
                    "Bump device": bump_device_mappings
                },
                'Default1': {
                    "Check list_Multi": check_list_mappings,
                    "Surface_Multi": surface_mappings,
                    "Pad device_Multi": pad_device_mappings,
                    "Bump device_Multi": bump_device_mappings
                }
            }

            # Track empty sheets
            empty_sheets = []

            for folder_type, mappings in all_mappings.items():
                for sheet_name, sheet_mappings in mappings.items():
                    if sheet_name in wb.sheetnames:
                        ws = wb[sheet_name]
                        for var, cell in sheet_mappings.items():
                            if var in self.variables.get(folder_type, {}) or var in self.variables:
                                value = self.variables.get(folder_type, {}).get(var) or self.variables.get(var)
                                try:
                                    value = float(value)
                                except ValueError:
                                    pass
                            
                                    ws[cell] = value
                            sheet_updated = True
                        
                        if not sheet_updated:
                            empty_sheets.append(sheet_name)

            def check_scan_area_ini(folder_type):
                ini_path = os.path.join(self.avi_recipe_path, 'Setup1', 'Recipes', folder_type, 'Zones', 'Scan Area.ini')
                print(f"Checking Scan Area.ini for {folder_type}: {ini_path}")
                if os.path.exists(ini_path):
                    config = configparser.ConfigParser()
                    config.read(ini_path)
                    enable_value = config.get('Surface', 'Enable', fallback='1')
                    print(f"Enable value for {folder_type}: {enable_value}")
                    return enable_value == '0'  # 如果 Enable 為 0，則返回 True（表示需要刪除工作表）
                print(f"Scan Area.ini not found for {folder_type}")
                return False  # 如果文件不存在，默認不刪除工作表

            # 在 update_excel_file 方法中
            default_should_delete = check_scan_area_ini('Default')
            default1_should_delete = check_scan_area_ini(self.default1_actual_name)

            print(f"Should delete Default Surface sheet: {default_should_delete}")
            print(f"Should delete Default1 Surface sheet: {default1_should_delete}")

            # Track empty sheets
            empty_sheets = []

            for folder_type, mappings in all_mappings.items():
                for sheet_name, sheet_mappings in mappings.items():
                    if sheet_name in wb.sheetnames:
                        ws = wb[sheet_name]
                        sheet_updated = False
                        for var, cell in sheet_mappings.items():
                            if var in self.variables.get(folder_type, {}):
                                value = self.variables[folder_type][var]
                            elif var in self.variables:
                                value = self.variables[var]
                            else:
                                continue
                            
                            try:
                                value = float(value)
                            except ValueError:
                                pass
                            print(f"Updating cell {cell} in sheet {sheet_name} with value {value}")
                            ws[cell] = value
                            sheet_updated = True
                        
                        if not sheet_updated:
                            empty_sheets.append(sheet_name)

            # Remove empty sheets
            for sheet_name in empty_sheets:
                if sheet_name in wb.sheetnames:
                    print(f"Removing empty sheet: {sheet_name}")
                    wb.remove(wb[sheet_name])

            # 更新 Excel 後，根據檢查結果刪除工作表
            if default_should_delete and 'Surface' in wb.sheetnames:
                wb.remove(wb['Surface'])
                print("工作表 'Surface' 已被刪除，因為 Default 的 Scan Area.ini 中 Enable=0")
            else:
                print("工作表 'Surface' 未被刪除")

            if default1_should_delete and 'Surface_Multi' in wb.sheetnames:
                wb.remove(wb['Surface_Multi'])
                print("工作表 'Surface_Multi' 已被刪除，因為 Default1 的 Scan Area.ini 中 Enable=0")
            else:
                print("工作表 'Surface_Multi' 未被刪除")

            for sheet_name in ["Pad device", "Bump device", "Pad device_Multi", "Bump device_Multi"]:
                if sheet_name in wb.sheetnames:
                    ws = wb[sheet_name]
                    for row in range(1, ws.max_row + 1): 
                        current_cell = ws.cell(row=row, column=6)  # Column F
                        
                        if row < ws.max_row:
                            next_cell = ws.cell(row=row+1, column=6)  # Next row's Column F
                        else:
                            next_cell = None  
                        
                        # Check if the cell is part of a merged range
                        is_merged = any(current_cell.coordinate in merged_range for merged_range in ws.merged_cells.ranges)
                        
                        if not is_merged:  # Only hide if not part of a merged cell
                            if current_cell.value is None:
                                # If F column is empty, hide the row
                                ws.row_dimensions[row].hidden = True
                            elif current_cell.value == "Setup File Value" and (next_cell is None or next_cell.value is None):
                                # If current cell is "Setup File Value" and next cell is empty or doesn't exist, hide the current row
                                ws.row_dimensions[row].hidden = True
                            else:
                                # Otherwise, make sure the row is visible
                                ws.row_dimensions[row].hidden = False
                        else:
                            # If the cell is part of a merged range, make sure the row is visible
                            ws.row_dimensions[row].hidden = False
                    
                    if sheet_name in ["Bump device", "Bump device_Multi"]:
                        if all(ws.row_dimensions[row].hidden for row in range(3, 59)):
                            ws.row_dimensions[2].hidden = True
                        
                        if all(ws.row_dimensions[row].hidden for row in range(61, 117)):
                            ws.row_dimensions[60].hidden = True
                        
                        if all(ws.row_dimensions[row].hidden for row in range(119, 175)):
                            ws.row_dimensions[118].hidden = True
                        
                        if all(ws.row_dimensions[row].hidden for row in range(177, 233)):
                            ws.row_dimensions[176].hidden = True
                        
                        if all(ws.row_dimensions[row].hidden for row in range(235, 291)):
                            ws.row_dimensions[234].hidden = True
                    
                    if sheet_name in ["Pad device", "Pad device_Multi"]:
                        if all(ws.row_dimensions[row].hidden for row in range(3, 25)):
                            ws.row_dimensions[2].hidden = True

                        if all(ws.row_dimensions[row].hidden for row in range(27, 49)):
                            ws.row_dimensions[26].hidden = True

                        if all(ws.row_dimensions[row].hidden for row in range(51, 73)):
                            ws.row_dimensions[50].hidden = True 

                        if all(ws.row_dimensions[row].hidden for row in range(75, 97)):
                            ws.row_dimensions[74].hidden = True 

                        if all(ws.row_dimensions[row].hidden for row in range(99, 121)):
                            ws.row_dimensions[98].hidden = True 

                        if all(ws.row_dimensions[row].hidden for row in range(124, 148)):
                            ws.row_dimensions[123].hidden = True 

                        if all(ws.row_dimensions[row].hidden for row in range(150, 174)):
                            ws.row_dimensions[149].hidden = True 

                        if all(ws.row_dimensions[row].hidden for row in range(176, 200)):
                            ws.row_dimensions[175].hidden = True 

                        if all(ws.row_dimensions[row].hidden for row in range(202, 226)):
                            ws.row_dimensions[201].hidden = True 

                        if all(ws.row_dimensions[row].hidden for row in range(228, 252)):
                            ws.row_dimensions[227].hidden = True 

                        if all(ws.row_dimensions[row].hidden for row in range(255, 282)):
                            ws.row_dimensions[254].hidden = True 

                        if all(ws.row_dimensions[row].hidden for row in range(284, 311)):
                            ws.row_dimensions[283].hidden = True 

                        if all(ws.row_dimensions[row].hidden for row in range(313, 340)):
                            ws.row_dimensions[312].hidden = True 

                        if all(ws.row_dimensions[row].hidden for row in range(342, 369)):
                            ws.row_dimensions[341].hidden = True 

                        if all(ws.row_dimensions[row].hidden for row in range(371, 398)):
                            ws.row_dimensions[370].hidden = True 

                else:
                    print(f"Sheet '{sheet_name}' not found, skipping.")

                sheets_to_check = ['Surface', 'Pad device', 'Bump device']
                multi_sheets_to_check = ['Surface_Multi', 'Pad device_Multi', 'Bump device_Multi']

                if all(sheet not in wb.sheetnames for sheet in sheets_to_check):
                    if 'Check list' in wb.sheetnames:
                        print("刪除 'Check list' 工作表，因為 'Surface'、'Pad device' 和 'Bump device' 都已被刪除")
                        wb.remove(wb['Check list'])

                if all(sheet not in wb.sheetnames for sheet in multi_sheets_to_check):
                    if 'Check list_Multi' in wb.sheetnames:
                        print("刪除 'Check list_Multi' 工作表，因為 'Surface_Multi'、'Pad device_Multi' 和 'Bump device_Multi' 都已被刪除")
                        wb.remove(wb['Check list_Multi'])

                # Protect updated cells and enable sheet protection
                updated_cells = set()

                for folder_type, mappings in all_mappings.items():
                    for sheet_name, sheet_mappings in mappings.items():
                        if sheet_name in wb.sheetnames:
                            ws = wb[sheet_name]
                            for var, cell in sheet_mappings.items():
                                if var in self.variables.get(folder_type, {}) or var in self.variables:
                                    value = self.variables.get(folder_type, {}).get(var) or self.variables.get(var)
                                    try:
                                        value = float(value)
                                    except ValueError:
                                        pass
                                    print(f"Updating cell {cell} in sheet {sheet_name} with value {value}")
                                    ws[cell] = value
                                    updated_cells.add((sheet_name, cell))

            # Protect updated cells and enable sheet protection
            for sheet_name, cell in updated_cells:
                ws = wb[sheet_name]
                ws[cell].protection = openpyxl.styles.Protection(locked=True)

            for ws in wb.worksheets:
                if ws.title not in ['Check list', 'Check list_Multi', 'Snapshot', 'Die shift check', 'Trial run']:
                    for row in ws.iter_rows(min_col=1, max_col=7):  # 從A列到G列
                        for cell in row:
                            cell.protection = openpyxl.styles.Protection(locked=True)
                    
                    # 對於這些工作表,其他列保持解鎖狀態
                    for row in ws.iter_rows(min_col=8): 
                        for cell in row:
                            cell.protection = openpyxl.styles.Protection(locked=False)

                else:
                    for row in ws.iter_rows():
                        for cell in row:
                            if (ws.title, cell.coordinate) not in updated_cells:
                                if ws.title in ['Check list', 'Check list_Multi']:
                                    if cell.column == 2:  # 這裡鎖定 B 列
                                        cell.protection = openpyxl.styles.Protection(locked=True)
                                    elif cell.coordinate in ['C26', 'C27', 'C28', 'C32', 'C33', 'C34', 'C40', 'C41', 'C49', 'C51', 'C52', 
                                                             'C53', 'C54', 'C55' , 'C58', 'C59', 'C62', 'C63', 'E4', 'F4', 'E5', 'F5', 'E7', 'F7', 'E8', 'F8',
                                                               'E16', 'F16', 'E17', 'F17', 'E18', 'F18', 'E19', 'F19', 'E20', 'F20', 'E21', 'F21', 'E23', 'F23',
                                                                 'E24', 'F24', 'E25', 'F25', 'E26', 'F26', 'E27', 'F27', 'E28', 'F28', 'E29', 'F29', 'E30' 'F30',
                                                                   'E31', 'F31', 'E35', 'F35', 'E36', 'F36', 'E37', 'F37', 'E38', 'F38', 'E42', 'F42', 'E43', 'F43',
                                                                     'E44', 'F44', 'E45', 'F45']:
                                        cell.protection = openpyxl.styles.Protection(locked=True)
                                    elif cell.column == 4 and cell.coordinate in ['D4', 'D5', 'D7', 'D8', 'D16', 'D17', 'D18', 'D19', 'D20', 'D21', 'D23',
                                                                                   'D24', 'D25', 'D26', 'D27', 'D28', 'D30', 'D31', 'D35', 'D36', 'D37', 'D38','D42', 'D43', 'D44', 'D45', 'D64']:
                                        cell.protection = openpyxl.styles.Protection(locked=True)
                                    else:
                                        cell.protection = openpyxl.styles.Protection(locked=False)
                                else:
                                    cell.protection = openpyxl.styles.Protection(locked=False)
                ws.protection.sheet = True
                ws.protection.password = 'Ardentec'
                ws.protection.enable()

            # Save the workbook after all updates
            wb.save(output_path)
            print(f"Excel file updated and protected successfully: {output_path}")
                        
        except Exception as e:
            print(f"An error occurred while updating the Excel file: {str(e)}")
            print("Traceback:")
            print(traceback.format_exc()) 
        finally:
            if 'wb' in locals():
                wb.close()

    def run(self):
        try:
            self.process_files()
            self.update_excel_file()
            
            print("Result：")
            print(json.dumps(self.variables, indent=2))
            
            self.processing_completed.emit()
        except Exception as e:
            error_message = str(e)
            if "Setup1\\Recipes\\file count >=" in error_message:
                message, path = error_message.split('|')
                self.error_occurred.emit(f"{message}\n點擊確定後將打開資料夾")
                self.open_folder_signal.emit(path)
            else:
                self.error_occurred.emit(str(e))

class AVIRecipeParser(QWidget):
    def __init__(self):
        super().__init__()
        self.initUI()
        self.check_version()
        self.save_log()

    def initUI(self):
        self.setWindowTitle('AVI Recipe check list')
        self.setGeometry(100, 100, 400, 250)
        self.center()

        layout = QVBoxLayout()

        self.select_button = QPushButton('選擇Recipe檔案')
        self.select_button.clicked.connect(self.select_recipe_folder)
        layout.addWidget(self.select_button)

        self.icon_label = QLabel()
        #icon_pixmap = QPixmap('format_1.ico').scaled(140, 140, Qt.KeepAspectRatio, Qt.SmoothTransformation)
        icon_pixmap = QPixmap(resource_path('format_1.ico')).scaled(140, 140, Qt.KeepAspectRatio, Qt.SmoothTransformation)
        self.icon_label.setPixmap(icon_pixmap)
        self.icon_label.setAlignment(Qt.AlignCenter)
        layout.addWidget(self.icon_label)

        self.generate_button = QPushButton('生成AVI check list')
        self.generate_button.clicked.connect(self.generate_check_list)
        self.generate_button.setEnabled(False)
        layout.addWidget(self.generate_button)

        self.progress_bar = QProgressBar()
        layout.addWidget(self.progress_bar)

        self.setLayout(layout)

    def center(self):
        qr = self.frameGeometry()
        cp = QDesktopWidget().availableGeometry().center()
        qr.moveCenter(cp)
        self.move(qr.topLeft())

    def select_recipe_folder(self):
        default_path = r"J:\Setupfile\Camtek\NPI"
        folder_path = QFileDialog.getExistingDirectory(self, "選擇Recipe檔案", default_path)
        if folder_path:
            self.avi_recipe_path = folder_path
            print(f"User selected path: {self.avi_recipe_path}")
            
            # 提取 'Recipe/' 之後的部分作為 AVI_recipe_name
            recipe_index = self.avi_recipe_path.rfind('Recipe/')
            if recipe_index != -1:
                self.AVI_recipe_name = self.avi_recipe_path[recipe_index + 7:]  # 7 是 'Recipe/' 的長度
            else:
                self.AVI_recipe_name = os.path.basename(self.avi_recipe_path)
            
            print(f"AVI_recipe_name: {self.AVI_recipe_name}")
            
            self.generate_button.setEnabled(True)
            self.update_icon('format_2.ico') 

    def update_icon(self, icon_file):
        icon_path = resource_path(icon_file)
        icon_pixmap = QPixmap(icon_path).scaled(145, 145, Qt.KeepAspectRatio, Qt.SmoothTransformation)
        self.icon_label.setPixmap(icon_pixmap)

    def generate_check_list(self):
        self.progress_bar.setValue(0)
        self.generate_button.setEnabled(False)
        self.select_button.setEnabled(False)

        try:
            self.file_processor = FileProcessor(self.avi_recipe_path)
            self.file_processor.progress_updated.connect(self.update_progress)
            self.file_processor.processing_completed.connect(self.processing_completed)
            self.file_processor.error_occurred.connect(self.show_error)
            self.file_processor.open_folder_signal.connect(self.open_folder)
            self.file_processor.start()
        except ValueError as e:
            self.show_error(str(e))
        except Exception as e:
            self.show_error(f"Wrong: {str(e)}")
        finally:
            self.generate_button.setEnabled(True)
            self.select_button.setEnabled(True)

    def update_progress(self, value):
        self.progress_bar.setValue(value)

    def processing_completed(self):
        self.progress_bar.setValue(100)
        reply = QMessageBox.information(self, "完成", "AVI check list 生成完成", QMessageBox.Ok)
        if reply == QMessageBox.Ok:
            self.open_output_file()
            self.close()

    def show_error(self, error_message):
        QMessageBox.critical(self, "錯誤", f"處理過程中發生錯誤：\n{error_message}")
        self.generate_button.setEnabled(True)
        self.select_button.setEnabled(True)

    def open_folder(self, path):
        os.startfile(path)

    def open_output_file(self):
        recipe_name = os.path.basename(self.avi_recipe_path)
        new_file_name = f"{recipe_name}_AVI check list.xlsx"
        output_path = os.path.join(os.path.expanduser("~"), "Downloads", new_file_name)
        if os.path.exists(output_path):
            os.startfile(output_path)
        else:
            QMessageBox.warning(self, "警告", f"無法找到文件: {new_file_name}")

    def save_log(self):
        try:
            hostname = socket.gethostname()
            match = re.search(r'^(.+)', hostname)
            username = match.group(1) if match else 'Unknown'

            current_datetime = datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S")
            log_folder = r'M:\QA_Program_Raw_Data\Log History'
            archive_path = os.path.join(log_folder, 'AVI Check list.7z')
            log_filename = f'{username}.txt'
            new_log_message = f"{current_datetime} {username} Open\n"
            os.makedirs(log_folder, exist_ok=True)

            if not os.path.exists(archive_path):
                with py7zr.SevenZipFile(archive_path, mode='w', password='@Joe11111111') as archive:
                    archive.writestr(new_log_message, f'AVI Check list/{log_filename}')
            else:
                log_content = ""
                files_to_keep = []

                with py7zr.SevenZipFile(archive_path, mode='r', password='@Joe11111111') as archive:
                    for filename, bio in archive.read().items():
                        if filename == f'AVI Check list/{log_filename}':
                            log_content = bio.read().decode('utf-8')
                        else:
                            files_to_keep.append((filename, bio.read()))

                if new_log_message not in log_content:
                    log_content += new_log_message

                with tempfile.NamedTemporaryFile(delete=False, suffix='.7z') as temp_file:
                    temp_archive_path = temp_file.name

                with py7zr.SevenZipFile(temp_archive_path, mode='w', password='@Joe11111111') as archive:
                    archive.writestr(log_content.encode('utf-8'), f'AVI Check list/{log_filename}')
                    for filename, content in files_to_keep:
                        archive.writestr(content, filename)

                shutil.move(temp_archive_path, archive_path)

        except Exception as e:
            print(f"寫入log時發生錯誤: {e}")

    def check_version(self):
        try:
            app_folder = r"M:\QA_Program_Raw_Data\Apps"
            exe_files = [f for f in os.listdir(app_folder) if f.startswith("AVI Check list_V") and f.endswith(".exe")]

            if not exe_files:
                QMessageBox.warning(self, '未獲取啟動權限', '未獲取啟動權限, 請申請M:\QA_Program_Raw_Data權限, 並聯絡#1082 Racky')
                sys.exit(1)

            # 修改版本號提取邏輯，只取主版本號
            latest_version = max(int(re.search(r'_V(\d+)', f).group(1)) for f in exe_files)

            # 修改當前版本號提取邏輯，只取主版本號
            current_version_match = re.search(r'_V(\d+)', os.path.basename(sys.executable))
            if current_version_match:
                current_version = int(current_version_match.group(1))
            else:
                current_version = 4

            if current_version < latest_version:
                QMessageBox.information(self, '請更新至最新版本', '請更新至最新版本')
                os.startfile(app_folder)  # 開啟指定的資料夾
                sys.exit(0)

            hostname = socket.gethostname()
            match = re.search(r'^(.+)', hostname)
            if match:
                username = match.group(1)
                if username == "A000000":
                    QMessageBox.warning(self, '未獲取啟動權限', '未獲取啟動權限, 請申請M:\QA_Program_Raw_Data權限, 並聯絡#1082 Racky')
                    sys.exit(1)
            else:
                QMessageBox.warning(self, '未獲取啟動權限', '未獲取啟動權限, 請申請M:\QA_Program_Raw_Data權限, 並聯絡#1082 Racky')
                sys.exit(1)

        except FileNotFoundError:
            QMessageBox.warning(self, '未獲取啟動權限', '未獲取啟動權限, 請申請M:\QA_Program_Raw_Data權限, 並聯絡#1082 Racky')
            sys.exit(1)
        
def get_application_path():
    if getattr(sys, 'frozen', False):
        return sys._MEIPASS
    else:
        return os.path.dirname(os.path.abspath(__file__))

if __name__ == '__main__':
    app = QApplication(sys.argv)
    application_path = get_application_path()
    icon_path = os.path.join(application_path, 'format.ico')
    app.setWindowIcon(QIcon(icon_path))

    font = QFont("微軟正黑體", 9)
    font.setBold(True)
    app.setFont(font)

    qtmodern.styles.dark(app)

    app.setFont(font)

    ex = AVIRecipeParser()
    ex.check_version()  
    win = qtmodern.windows.ModernWindow(ex)
    win.show()

    sys.exit(app.exec_())